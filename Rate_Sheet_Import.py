import os, sys
os.environ.setdefault('DJANGO_SETTINGS_MODULE', 'VerusPricer.settings')
import django
django.setup()
from pricer.models import BaseRate, GradeAdj, DocTypeAdj, DTIAdj, LoanBalanceAdj, PurposeAdj, OccupancyAdj, PropertyTypeAdj, IOAdj, LoanTermAdj, StateAdj, PrepayAdj, ReservesAdj
from datetime import datetime
from openpyxl import *

def importRateSheets():
	#Reset all data models for clean refresh of rate sheets
	print "Clearing Exisiting Models..."
	BaseRate.objects.all().delete()
	GradeAdj.objects.all().delete()
	DocTypeAdj.objects.all().delete()
	DTIAdj.objects.all().delete()
	PurposeAdj.objects.all().delete()
	LoanBalanceAdj.objects.all().delete()
	PropertyTypeAdj.objects.all().delete()
	StateAdj.objects.all().delete()
	IOAdj.objects.all().delete()
	PrepayAdj.objects.all().delete()
	LoanTermAdj.objects.all().delete()
	OccupancyAdj.objects.all().delete()
	ReservesAdj.objects.all().delete()

	# CREDIT ASCENT RATE SHEET
	print "Loading Verus Credit Ascent Rate Sheet..."
	wb = load_workbook(filename = '/home/invictus/VerusPricer/static/rates/Verus Credit Ascent Rate Sheet.xlsx')
	ws = wb.active
	
	#Load Base Rates
	startRow = 8
	endRow = 20
	startCol = 2
	endCol = 9
	for row in range(startRow, endRow+1):
		for col in range(startCol, endCol+1):
			if ws.cell(row=row, column=col).value:
				if ws.cell(row=row, column=startCol-1).value == "720+":
					rate = BaseRate.objects.get_or_create(rateSheetName = "Credit Ascent", fico = "720-739", ltv = ws.cell(row=startRow-1, column=col).value, adj = ws.cell(row=row, column=col).value)
					rate = BaseRate.objects.get_or_create(rateSheetName = "Credit Ascent", fico = "740+", ltv = ws.cell(row=startRow-1, column=col).value, adj = ws.cell(row=row, column=col).value)
				else:
					rate = BaseRate.objects.get_or_create(rateSheetName = "Credit Ascent", fico = ws.cell(row=row, column=startCol-1).value, ltv = ws.cell(row=startRow-1, column=col).value, adj = ws.cell(row=row, column=col).value)
				#rate.save()

	#Load Grade Adjustments
	startRow = 24
	endRow = 29
	startCol = 2
	endCol = 9
	for row in range(startRow, endRow+1):
		for col in range(startCol, endCol+1):
			if ws.cell(row=row, column=col).value != "NA":
				adj = GradeAdj.objects.get_or_create(rateSheetName = "Credit Ascent", grade = ws.cell(row=row, column=startCol-1).value, ltv = ws.cell(row=startRow-1, column=col).value, adj = ws.cell(row=row, column=col).value)
	for col in range(startCol, endCol+1):
		dj = GradeAdj.objects.get_or_create(rateSheetName = "Credit Ascent", grade = "Foreign Credit", ltv = ws.cell(row=startRow-1, column=col).value, adj = 0)

	#Load DocType Adjustments
	startRow = 31
	endRow = 34
	startCol = 2
	endCol = 9
	for row in range(startRow, endRow+1):
		for col in range(startCol, endCol+1):
			if ws.cell(row=row, column=col).value != "NA":
				if ws.cell(row=row, column=startCol-1).value == "24 Mo Bank Statement: A, A- Grades":
					adj = DocTypeAdj.objects.get_or_create(rateSheetName = "Credit Ascent", docType = "24 Mo Bank Statement", grade = "A", ltv = ws.cell(row=startRow-1, column=col).value, adj = ws.cell(row=row, column=col).value)
					adj = DocTypeAdj.objects.get_or_create(rateSheetName = "Credit Ascent", docType = "24 Mo Bank Statement", grade = "A-", ltv = ws.cell(row=startRow-1, column=col).value, adj = ws.cell(row=row, column=col).value)
				elif ws.cell(row=row, column=startCol-1).value == "24 Mo Bank Statement: B+, B, B- ,C Grades":
					adj = DocTypeAdj.objects.get_or_create(rateSheetName = "Credit Ascent", docType = "24 Mo Bank Statement", grade = "B+", ltv = ws.cell(row=startRow-1, column=col).value, adj = ws.cell(row=row, column=col).value)
					adj = DocTypeAdj.objects.get_or_create(rateSheetName = "Credit Ascent", docType = "24 Mo Bank Statement", grade = "B", ltv = ws.cell(row=startRow-1, column=col).value, adj = ws.cell(row=row, column=col).value)
					adj = DocTypeAdj.objects.get_or_create(rateSheetName = "Credit Ascent", docType = "24 Mo Bank Statement", grade = "B-", ltv = ws.cell(row=startRow-1, column=col).value, adj = ws.cell(row=row, column=col).value)
					adj = DocTypeAdj.objects.get_or_create(rateSheetName = "Credit Ascent", docType = "24 Mo Bank Statement", grade = "C", ltv = ws.cell(row=startRow-1, column=col).value, adj = ws.cell(row=row, column=col).value)
				else:
					adj = DocTypeAdj.objects.get_or_create(rateSheetName = "Credit Ascent", docType = ws.cell(row=row, column=startCol-1).value, grade = "A", ltv = ws.cell(row=startRow-1, column=col).value, adj = ws.cell(row=row, column=col).value)
					adj = DocTypeAdj.objects.get_or_create(rateSheetName = "Credit Ascent", docType = ws.cell(row=row, column=startCol-1).value, grade = "A-", ltv = ws.cell(row=startRow-1, column=col).value, adj = ws.cell(row=row, column=col).value)
					adj = DocTypeAdj.objects.get_or_create(rateSheetName = "Credit Ascent", docType = ws.cell(row=row, column=startCol-1).value, grade = "B+", ltv = ws.cell(row=startRow-1, column=col).value, adj = ws.cell(row=row, column=col).value)
					adj = DocTypeAdj.objects.get_or_create(rateSheetName = "Credit Ascent", docType = ws.cell(row=row, column=startCol-1).value, grade = "B", ltv = ws.cell(row=startRow-1, column=col).value, adj = ws.cell(row=row, column=col).value)
					adj = DocTypeAdj.objects.get_or_create(rateSheetName = "Credit Ascent", docType = ws.cell(row=row, column=startCol-1).value, grade = "B-", ltv = ws.cell(row=startRow-1, column=col).value, adj = ws.cell(row=row, column=col).value)
					adj = DocTypeAdj.objects.get_or_create(rateSheetName = "Credit Ascent", docType = ws.cell(row=row, column=startCol-1).value, grade = "C", ltv = ws.cell(row=startRow-1, column=col).value, adj = ws.cell(row=row, column=col).value)
	for col in range(startCol, endCol+1):
		adj = DocTypeAdj.objects.get_or_create(rateSheetName = "Credit Ascent", docType = "Full Doc", grade = "A", ltv = ws.cell(row=startRow-1, column=col).value, adj = 0.0)
		adj = DocTypeAdj.objects.get_or_create(rateSheetName = "Credit Ascent", docType = "Full Doc", grade = "A-", ltv = ws.cell(row=startRow-1, column=col).value, adj = 0.0)
		adj = DocTypeAdj.objects.get_or_create(rateSheetName = "Credit Ascent", docType = "Full Doc", grade = "B+", ltv = ws.cell(row=startRow-1, column=col).value, adj = 0.0)
		adj = DocTypeAdj.objects.get_or_create(rateSheetName = "Credit Ascent", docType = "Full Doc", grade = "B", ltv = ws.cell(row=startRow-1, column=col).value, adj = 0.0)
		adj = DocTypeAdj.objects.get_or_create(rateSheetName = "Credit Ascent", docType = "Full Doc", grade = "B-", ltv = ws.cell(row=startRow-1, column=col).value, adj = 0.0)
		adj = DocTypeAdj.objects.get_or_create(rateSheetName = "Credit Ascent", docType = "Full Doc", grade = "C", ltv = ws.cell(row=startRow-1, column=col).value, adj = 0.0)

	#Load DTI Adjustments
	startRow = 36
	endRow = 37
	startCol = 2
	endCol = 9
	for row in range(startRow, endRow+1):
		for col in range(startCol, endCol+1):
			if ws.cell(row=row, column=col).value != "NA":
				if ws.cell(row=row, column=startCol-1).value == "> 43%: A, A- Grades":
					adj = DTIAdj.objects.get_or_create(rateSheetName = "Credit Ascent", dti = "43.01% - 50%", grade = "A", ltv = ws.cell(row=startRow-1, column=col).value, adj = ws.cell(row=row, column=col).value)
					adj = DTIAdj.objects.get_or_create(rateSheetName = "Credit Ascent", dti = "43.01% - 50%", grade = "A-", ltv = ws.cell(row=startRow-1, column=col).value, adj = ws.cell(row=row, column=col).value)
					adj = DTIAdj.objects.get_or_create(rateSheetName = "Credit Ascent", dti = "> 50%", grade = "A", ltv = ws.cell(row=startRow-1, column=col).value, adj = ws.cell(row=row, column=col).value)
					adj = DTIAdj.objects.get_or_create(rateSheetName = "Credit Ascent", dti = "> 50%", grade = "A-", ltv = ws.cell(row=startRow-1, column=col).value, adj = ws.cell(row=row, column=col).value)
				elif ws.cell(row=row, column=startCol-1).value == "> 43%: B+, B, B- ,C Grades & Foreign National":
					adj = DTIAdj.objects.get_or_create(rateSheetName = "Credit Ascent", dti = "43.01% - 50%", grade = "B+", ltv = ws.cell(row=startRow-1, column=col).value, adj = ws.cell(row=row, column=col).value)
					adj = DTIAdj.objects.get_or_create(rateSheetName = "Credit Ascent", dti = "43.01% - 50%", grade = "B", ltv = ws.cell(row=startRow-1, column=col).value, adj = ws.cell(row=row, column=col).value)
					adj = DTIAdj.objects.get_or_create(rateSheetName = "Credit Ascent", dti = "43.01% - 50%", grade = "B-", ltv = ws.cell(row=startRow-1, column=col).value, adj = ws.cell(row=row, column=col).value)
					adj = DTIAdj.objects.get_or_create(rateSheetName = "Credit Ascent", dti = "43.01% - 50%", grade = "C", ltv = ws.cell(row=startRow-1, column=col).value, adj = ws.cell(row=row, column=col).value)
					adj = DTIAdj.objects.get_or_create(rateSheetName = "Credit Ascent", dti = "> 50%", grade = "B+", ltv = ws.cell(row=startRow-1, column=col).value, adj = ws.cell(row=row, column=col).value)
					adj = DTIAdj.objects.get_or_create(rateSheetName = "Credit Ascent", dti = "> 50%", grade = "B", ltv = ws.cell(row=startRow-1, column=col).value, adj = ws.cell(row=row, column=col).value)
					adj = DTIAdj.objects.get_or_create(rateSheetName = "Credit Ascent", dti = "> 50%", grade = "B-", ltv = ws.cell(row=startRow-1, column=col).value, adj = ws.cell(row=row, column=col).value)
					adj = DTIAdj.objects.get_or_create(rateSheetName = "Credit Ascent", dti = "> 50%", grade = "C", ltv = ws.cell(row=startRow-1, column=col).value, adj = ws.cell(row=row, column=col).value)
				else:
					adj = DTIAdj.objects.get_or_create(rateSheetName = "Credit Ascent", dti = ws.cell(row=row, column=startCol-1).value, ltv = ws.cell(row=startRow-1, column=col).value, adj = ws.cell(row=row, column=col).value)
	for col in range(startCol, endCol+1):
		adj = DTIAdj.objects.get_or_create(rateSheetName = "Credit Ascent", dti = "<= 36%", grade = "A", ltv = ws.cell(row=startRow-1, column=col).value, adj = 0.0)
		adj = DTIAdj.objects.get_or_create(rateSheetName = "Credit Ascent", dti = "<= 36%", grade = "A-", ltv = ws.cell(row=startRow-1, column=col).value, adj = 0.0)
		adj = DTIAdj.objects.get_or_create(rateSheetName = "Credit Ascent", dti = "<= 36%", grade = "B+", ltv = ws.cell(row=startRow-1, column=col).value, adj = 0.0)
		adj = DTIAdj.objects.get_or_create(rateSheetName = "Credit Ascent", dti = "<= 36%", grade = "B", ltv = ws.cell(row=startRow-1, column=col).value, adj = 0.0)
		adj = DTIAdj.objects.get_or_create(rateSheetName = "Credit Ascent", dti = "<= 36%", grade = "B-", ltv = ws.cell(row=startRow-1, column=col).value, adj = 0.0)
		adj = DTIAdj.objects.get_or_create(rateSheetName = "Credit Ascent", dti = "<= 36%", grade = "C", ltv = ws.cell(row=startRow-1, column=col).value, adj = 0.0)
		adj = DTIAdj.objects.get_or_create(rateSheetName = "Credit Ascent", dti = "36.01% - 43%", grade = "A", ltv = ws.cell(row=startRow-1, column=col).value, adj = 0.0)
		adj = DTIAdj.objects.get_or_create(rateSheetName = "Credit Ascent", dti = "36.01% - 43%", grade = "A-", ltv = ws.cell(row=startRow-1, column=col).value, adj = 0.0)
		adj = DTIAdj.objects.get_or_create(rateSheetName = "Credit Ascent", dti = "36.01% - 43%", grade = "B+", ltv = ws.cell(row=startRow-1, column=col).value, adj = 0.0)
		adj = DTIAdj.objects.get_or_create(rateSheetName = "Credit Ascent", dti = "36.01% - 43%", grade = "B", ltv = ws.cell(row=startRow-1, column=col).value, adj = 0.0)
		adj = DTIAdj.objects.get_or_create(rateSheetName = "Credit Ascent", dti = "36.01% - 43%", grade = "B-", ltv = ws.cell(row=startRow-1, column=col).value, adj = 0.0)
		adj = DTIAdj.objects.get_or_create(rateSheetName = "Credit Ascent", dti = "36.01% - 43%", grade = "C", ltv = ws.cell(row=startRow-1, column=col).value, adj = 0.0)

	#Load Loan Balance Adjustments
	startRow = 39
	endRow = 42
	startCol = 2
	endCol = 9
	for row in range(startRow, endRow+1):
		for col in range(startCol, endCol+1):
			if ws.cell(row=row, column=col).value != "NA":
				if ws.cell(row=row, column=startCol-1).value == "<150k":
					adj = LoanBalanceAdj.objects.get_or_create(rateSheetName = "Credit Ascent", balance = "$100,000 - $149,999", ltv = ws.cell(row=startRow-1, column=col).value, adj = ws.cell(row=row, column=col).value)
					adj = LoanBalanceAdj.objects.get_or_create(rateSheetName = "Credit Ascent", balance = "< $100,000", ltv = ws.cell(row=startRow-1, column=col).value, adj = ws.cell(row=row, column=col).value)
				elif ws.cell(row=row, column=startCol-1).value == "750k to $1mm":
					adj = LoanBalanceAdj.objects.get_or_create(rateSheetName = "Credit Ascent", balance = "$750,000 - $799,999", ltv = ws.cell(row=startRow-1, column=col).value, adj = ws.cell(row=row, column=col).value)
					adj = LoanBalanceAdj.objects.get_or_create(rateSheetName = "Credit Ascent", balance = "$800,000 - $999,999", ltv = ws.cell(row=startRow-1, column=col).value, adj = ws.cell(row=row, column=col).value)
				elif ws.cell(row=row, column=startCol-1).value == "$1mm to $1.5mm":
					adj = LoanBalanceAdj.objects.get_or_create(rateSheetName = "Credit Ascent", balance = "$1,000,000 - $1,249,999", ltv = ws.cell(row=startRow-1, column=col).value, adj = ws.cell(row=row, column=col).value)
					adj = LoanBalanceAdj.objects.get_or_create(rateSheetName = "Credit Ascent", balance = "$1,250,000 - $1,499,999", ltv = ws.cell(row=startRow-1, column=col).value, adj = ws.cell(row=row, column=col).value)
				elif ws.cell(row=row, column=startCol-1).value == ">= $1.5mm":
					adj = LoanBalanceAdj.objects.get_or_create(rateSheetName = "Credit Ascent", balance = "$1,500,000 - $2,000,000", ltv = ws.cell(row=startRow-1, column=col).value, adj = ws.cell(row=row, column=col).value)
				else:
					adj = LoanBalanceAdj.objects.get_or_create(rateSheetName = "Credit Ascent", balance = ws.cell(row=row, column=startCol-1).value, ltv = ws.cell(row=startRow-1, column=col).value, adj = ws.cell(row=row, column=col).value)
	for col in range(startCol, endCol+1):
		adj = LoanBalanceAdj.objects.get_or_create(rateSheetName = "Credit Ascent", balance = "$150,000 - $399,999", ltv = ws.cell(row=startRow-1, column=col).value, adj = 0.0)
		adj = LoanBalanceAdj.objects.get_or_create(rateSheetName = "Credit Ascent", balance = "$400,000 - $499,999", ltv = ws.cell(row=startRow-1, column=col).value, adj = 0.0)
		adj = LoanBalanceAdj.objects.get_or_create(rateSheetName = "Credit Ascent", balance = "$500,000 - $749,999", ltv = ws.cell(row=startRow-1, column=col).value, adj = 0.0)

	#Load Purpose Adjustments
	startRow = 44
	endRow = 45
	startCol = 2
	endCol = 9
	for row in range(startRow, endRow+1):
		for col in range(startCol, endCol+1):
			if ws.cell(row=row, column=col).value != "NA":
				if ws.cell(row=row, column=startCol-1).value == "Cash Out: A, A- Grades":
					adj = PurposeAdj.objects.get_or_create(rateSheetName = "Credit Ascent", purpose = "Cash Out Refinance", grade = "A", ltv = ws.cell(row=startRow-1, column=col).value, adj = ws.cell(row=row, column=col).value)
					adj = PurposeAdj.objects.get_or_create(rateSheetName = "Credit Ascent", purpose = "Cash Out Refinance", grade = "A-", ltv = ws.cell(row=startRow-1, column=col).value, adj = ws.cell(row=row, column=col).value)
				elif ws.cell(row=row, column=startCol-1).value == "Cash Out: B+, B, B- ,C Grades":
					adj = PurposeAdj.objects.get_or_create(rateSheetName = "Credit Ascent", purpose = "Cash Out Refinance", grade = "B+", ltv = ws.cell(row=startRow-1, column=col).value, adj = ws.cell(row=row, column=col).value)			
					adj = PurposeAdj.objects.get_or_create(rateSheetName = "Credit Ascent", purpose = "Cash Out Refinance", grade = "B", ltv = ws.cell(row=startRow-1, column=col).value, adj = ws.cell(row=row, column=col).value)
					adj = PurposeAdj.objects.get_or_create(rateSheetName = "Credit Ascent", purpose = "Cash Out Refinance", grade = "B-", ltv = ws.cell(row=startRow-1, column=col).value, adj = ws.cell(row=row, column=col).value)
					adj = PurposeAdj.objects.get_or_create(rateSheetName = "Credit Ascent", purpose = "Cash Out Refinance", grade = "C", ltv = ws.cell(row=startRow-1, column=col).value, adj = ws.cell(row=row, column=col).value)
	for col in range(startCol, endCol+1):
		adj = PurposeAdj.objects.get_or_create(rateSheetName = "Credit Ascent", purpose = "Purchase", grade = "A", ltv = ws.cell(row=startRow-1, column=col).value, adj = 0.0)
		adj = PurposeAdj.objects.get_or_create(rateSheetName = "Credit Ascent", purpose = "Purchase", grade = "A-", ltv = ws.cell(row=startRow-1, column=col).value, adj = 0.0)
		adj = PurposeAdj.objects.get_or_create(rateSheetName = "Credit Ascent", purpose = "Purchase", grade = "B+", ltv = ws.cell(row=startRow-1, column=col).value, adj = 0.0)
		adj = PurposeAdj.objects.get_or_create(rateSheetName = "Credit Ascent", purpose = "Purchase", grade = "B", ltv = ws.cell(row=startRow-1, column=col).value, adj = 0.0)
		adj = PurposeAdj.objects.get_or_create(rateSheetName = "Credit Ascent", purpose = "Purchase", grade = "B-", ltv = ws.cell(row=startRow-1, column=col).value, adj = 0.0)
		adj = PurposeAdj.objects.get_or_create(rateSheetName = "Credit Ascent", purpose = "Purchase", grade = "C", ltv = ws.cell(row=startRow-1, column=col).value, adj = 0.0)
		adj = PurposeAdj.objects.get_or_create(rateSheetName = "Credit Ascent", purpose = "Rate/Term Refinance", grade = "A", ltv = ws.cell(row=startRow-1, column=col).value, adj = 0.0)
		adj = PurposeAdj.objects.get_or_create(rateSheetName = "Credit Ascent", purpose = "Rate/Term Refinance", grade = "A-", ltv = ws.cell(row=startRow-1, column=col).value, adj = 0.0)
		adj = PurposeAdj.objects.get_or_create(rateSheetName = "Credit Ascent", purpose = "Rate/Term Refinance", grade = "B+", ltv = ws.cell(row=startRow-1, column=col).value, adj = 0.0)
		adj = PurposeAdj.objects.get_or_create(rateSheetName = "Credit Ascent", purpose = "Rate/Term Refinance", grade = "B", ltv = ws.cell(row=startRow-1, column=col).value, adj = 0.0)
		adj = PurposeAdj.objects.get_or_create(rateSheetName = "Credit Ascent", purpose = "Rate/Term Refinance", grade = "B-", ltv = ws.cell(row=startRow-1, column=col).value, adj = 0.0)
		adj = PurposeAdj.objects.get_or_create(rateSheetName = "Credit Ascent", purpose = "Rate/Term Refinance", grade = "C", ltv = ws.cell(row=startRow-1, column=col).value, adj = 0.0)

	#Load Occupancy Adjustments
	startRow = 47
	endRow = 47
	startCol = 2
	endCol = 9
	for row in range(startRow, endRow+1):
		for col in range(startCol, endCol+1):
			if ws.cell(row=row, column=col).value != "NA":
				adj = OccupancyAdj.objects.get_or_create(rateSheetName = "Credit Ascent", occupancy = ws.cell(row=row, column=startCol-1).value, ltv = ws.cell(row=startRow-1, column=col).value, adj = ws.cell(row=row, column=col).value)
	for col in range(startCol, endCol+1):
		adj = OccupancyAdj.objects.get_or_create(rateSheetName = "Credit Ascent", occupancy = "Primary", ltv = ws.cell(row=startRow-1, column=col).value, adj = 0.0)

	#Load Property Type Adjustments
	startRow = 49
	endRow = 51
	startCol = 2
	endCol = 9
	for row in range(startRow, endRow+1):
		for col in range(startCol, endCol+1):
			if ws.cell(row=row, column=col).value != "NA":
				adj = PropertyTypeAdj.objects.get_or_create(rateSheetName = "Credit Ascent", propertyType = ws.cell(row=row, column=startCol-1).value, ltv = ws.cell(row=startRow-1, column=col).value, adj = ws.cell(row=row, column=col).value)
	for col in range(startCol, endCol+1):
		adj = PropertyTypeAdj.objects.get_or_create(rateSheetName = "Credit Ascent", propertyType = "SFR", ltv = ws.cell(row=startRow-1, column=col).value, adj = 0.0)
		adj = PropertyTypeAdj.objects.get_or_create(rateSheetName = "Credit Ascent", propertyType = "PUD", ltv = ws.cell(row=startRow-1, column=col).value, adj = 0.0)

	#Load State Adjustments
	startRow = 53
	endRow = 55
	startCol = 2
	endCol = 9
	for row in range(startRow, endRow+1):
		for col in range(startCol, endCol+1):
			if ws.cell(row=row, column=col).value != "NA":
				adj = StateAdj.objects.get_or_create(rateSheetName = "Credit Ascent", state = ws.cell(row=row, column=startCol-1).value, ltv = ws.cell(row=startRow-1, column=col).value, adj = ws.cell(row=row, column=col).value)

	#Load Interest Only Adjustments
	startRow = 57
	endRow = 57
	startCol = 2
	endCol = 9
	for row in range(startRow, endRow+1):
		for col in range(startCol, endCol+1):
			if ws.cell(row=row, column=col).value != "NA":
				adj = IOAdj.objects.get_or_create(rateSheetName = "Credit Ascent", io = "Yes", ltv = ws.cell(row=startRow-1, column=col).value, adj = ws.cell(row=row, column=col).value)
	for col in range(startCol, endCol+1):
		adj = IOAdj.objects.get_or_create(rateSheetName = "Credit Ascent", io = "No", ltv = ws.cell(row=startRow-1, column=col).value, adj = 0.0)

	#Load Loan Term Adjustments
	adj = LoanTermAdj.objects.get_or_create(rateSheetName = "Credit Ascent", loanTerm = "5/1 ARM" , adj = 0.0)
	adj = LoanTermAdj.objects.get_or_create(rateSheetName = "Credit Ascent", loanTerm = "7/1 ARM" , adj = 0.1)
	adj = LoanTermAdj.objects.get_or_create(rateSheetName = "Credit Ascent", loanTerm = "15 Yr Fix" , adj = 0.1)
	adj = LoanTermAdj.objects.get_or_create(rateSheetName = "Credit Ascent", loanTerm = "30 Yr Fix" , adj = 0.3)



	# INVESTOR DTI RATE SHEET
	print "Loading Verus Investor DTI Rate Sheet..."
	wb = load_workbook(filename = '/home/invictus/VerusPricer/static/rates/Verus Investor DTI Rate Sheet.xlsx')
	ws = wb.active
	
	#Load Base Rates
	startRow = 9
	endRow = 19
	startCol = 3
	endCol = 8
	for row in range(startRow, endRow+1):
		for col in range(startCol, endCol+1):
			if ws.cell(row=row, column=col).value:
				if ws.cell(row=row, column=startCol-2).value == "720+": 
					rate = BaseRate.objects.get_or_create(rateSheetName = "Investor DTI", fico = "720-739", ltv = ws.cell(row=startRow-1, column=col).value, adj = ws.cell(row=row, column=col).value)
					rate = BaseRate.objects.get_or_create(rateSheetName = "Investor DTI", fico = "740+", ltv = ws.cell(row=startRow-1, column=col).value, adj = ws.cell(row=row, column=col).value)
				else:
					rate = BaseRate.objects.get_or_create(rateSheetName = "Investor DTI", fico = ws.cell(row=row, column=startCol-2).value, ltv = ws.cell(row=startRow-1, column=col).value, adj = ws.cell(row=row, column=col).value)
				#rate.save()

	#Load Grade Adjustments
	startRow = 23
	endRow = 27
	startCol = 3
	endCol = 8
	for row in range(startRow, endRow+1):
		for col in range(startCol, endCol+1):
			if ws.cell(row=row, column=col).value != "NA":
				adj = GradeAdj.objects.get_or_create(rateSheetName = "Investor DTI", grade = ws.cell(row=row, column=startCol-2).value, ltv = ws.cell(row=startRow-1, column=col).value, adj = ws.cell(row=row, column=col).value)
	for col in range(startCol, endCol+1):
		dj = GradeAdj.objects.get_or_create(rateSheetName = "Investor DTI", grade = "Foreign Credit", ltv = ws.cell(row=startRow-1, column=col).value, adj = 0)

	#Load DocType Adjustments
	startRow = 29
	endRow = 32
	startCol = 3
	endCol = 8
	for row in range(startRow, endRow+1):
		for col in range(startCol, endCol+1):
			if ws.cell(row=row, column=col).value != "NA":
				if ws.cell(row=row, column=startCol-2).value == "24 Mo Bank Statement: A, A- Grades":
					adj = DocTypeAdj.objects.get_or_create(rateSheetName = "Investor DTI", docType = "24 Mo Bank Statement", grade = "A", ltv = ws.cell(row=startRow-1, column=col).value, adj = ws.cell(row=row, column=col).value)
					adj = DocTypeAdj.objects.get_or_create(rateSheetName = "Investor DTI", docType = "24 Mo Bank Statement", grade = "A-", ltv = ws.cell(row=startRow-1, column=col).value, adj = ws.cell(row=row, column=col).value)
				elif ws.cell(row=row, column=startCol-2).value == "24 Mo Bank Statement: B+, B, B- Grades":
					adj = DocTypeAdj.objects.get_or_create(rateSheetName = "Investor DTI", docType = "24 Mo Bank Statement", grade = "B+", ltv = ws.cell(row=startRow-1, column=col).value, adj = ws.cell(row=row, column=col).value)
					adj = DocTypeAdj.objects.get_or_create(rateSheetName = "Investor DTI", docType = "24 Mo Bank Statement", grade = "B", ltv = ws.cell(row=startRow-1, column=col).value, adj = ws.cell(row=row, column=col).value)
					adj = DocTypeAdj.objects.get_or_create(rateSheetName = "Investor DTI", docType = "24 Mo Bank Statement", grade = "B-", ltv = ws.cell(row=startRow-1, column=col).value, adj = ws.cell(row=row, column=col).value)
				else:
					adj = DocTypeAdj.objects.get_or_create(rateSheetName = "Investor DTI", docType = ws.cell(row=row, column=startCol-2).value, grade = "A", ltv = ws.cell(row=startRow-1, column=col).value, adj = ws.cell(row=row, column=col).value)
					adj = DocTypeAdj.objects.get_or_create(rateSheetName = "Investor DTI", docType = ws.cell(row=row, column=startCol-2).value, grade = "A-", ltv = ws.cell(row=startRow-1, column=col).value, adj = ws.cell(row=row, column=col).value)
					adj = DocTypeAdj.objects.get_or_create(rateSheetName = "Investor DTI", docType = ws.cell(row=row, column=startCol-2).value, grade = "B+", ltv = ws.cell(row=startRow-1, column=col).value, adj = ws.cell(row=row, column=col).value)
					adj = DocTypeAdj.objects.get_or_create(rateSheetName = "Investor DTI", docType = ws.cell(row=row, column=startCol-2).value, grade = "B", ltv = ws.cell(row=startRow-1, column=col).value, adj = ws.cell(row=row, column=col).value)
					adj = DocTypeAdj.objects.get_or_create(rateSheetName = "Investor DTI", docType = ws.cell(row=row, column=startCol-2).value, grade = "B-", ltv = ws.cell(row=startRow-1, column=col).value, adj = ws.cell(row=row, column=col).value)

	for col in range(startCol, endCol+1):
		adj = DocTypeAdj.objects.get_or_create(rateSheetName = "Investor DTI", docType = "Full Doc", grade = "A", ltv = ws.cell(row=startRow-1, column=col).value, adj = 0.0)
		adj = DocTypeAdj.objects.get_or_create(rateSheetName = "Investor DTI", docType = "Full Doc", grade = "A-", ltv = ws.cell(row=startRow-1, column=col).value, adj = 0.0)
		adj = DocTypeAdj.objects.get_or_create(rateSheetName = "Investor DTI", docType = "Full Doc", grade = "B+", ltv = ws.cell(row=startRow-1, column=col).value, adj = 0.0)
		adj = DocTypeAdj.objects.get_or_create(rateSheetName = "Investor DTI", docType = "Full Doc", grade = "B", ltv = ws.cell(row=startRow-1, column=col).value, adj = 0.0)
		adj = DocTypeAdj.objects.get_or_create(rateSheetName = "Investor DTI", docType = "Full Doc", grade = "B-", ltv = ws.cell(row=startRow-1, column=col).value, adj = 0.0)

	#Load DTI Adjustments 
	startRow = 34
	endRow = 35
	startCol = 3
	endCol = 8
	for row in range(startRow, endRow+1):
		for col in range(startCol, endCol+1):
			if ws.cell(row=row, column=col).value != "NA":
				if ws.cell(row=row, column=startCol-2).value == "> 43%: A, A- Grades":
					adj = DTIAdj.objects.get_or_create(rateSheetName = "Investor DTI", dti = "43.01% - 50%", grade = "A", ltv = ws.cell(row=startRow-1, column=col).value, adj = ws.cell(row=row, column=col).value)
					adj = DTIAdj.objects.get_or_create(rateSheetName = "Investor DTI", dti = "43.01% - 50%", grade = "A-", ltv = ws.cell(row=startRow-1, column=col).value, adj = ws.cell(row=row, column=col).value)
					adj = DTIAdj.objects.get_or_create(rateSheetName = "Investor DTI", dti = "> 50%", grade = "A", ltv = ws.cell(row=startRow-1, column=col).value, adj = ws.cell(row=row, column=col).value)
					adj = DTIAdj.objects.get_or_create(rateSheetName = "Investor DTI", dti = "> 50%", grade = "A-", ltv = ws.cell(row=startRow-1, column=col).value, adj = ws.cell(row=row, column=col).value)
				elif ws.cell(row=row, column=startCol-2).value == "> 43%: B+, B, B- Grades":
					adj = DTIAdj.objects.get_or_create(rateSheetName = "Investor DTI", dti = "43.01% - 50%", grade = "B+", ltv = ws.cell(row=startRow-1, column=col).value, adj = ws.cell(row=row, column=col).value)
					adj = DTIAdj.objects.get_or_create(rateSheetName = "Investor DTI", dti = "43.01% - 50%", grade = "B", ltv = ws.cell(row=startRow-1, column=col).value, adj = ws.cell(row=row, column=col).value)
					adj = DTIAdj.objects.get_or_create(rateSheetName = "Investor DTI", dti = "43.01% - 50%", grade = "B-", ltv = ws.cell(row=startRow-1, column=col).value, adj = ws.cell(row=row, column=col).value)
					adj = DTIAdj.objects.get_or_create(rateSheetName = "Investor DTI", dti = "> 50%", grade = "B+", ltv = ws.cell(row=startRow-1, column=col).value, adj = ws.cell(row=row, column=col).value)
					adj = DTIAdj.objects.get_or_create(rateSheetName = "Investor DTI", dti = "> 50%", grade = "B", ltv = ws.cell(row=startRow-1, column=col).value, adj = ws.cell(row=row, column=col).value)
					adj = DTIAdj.objects.get_or_create(rateSheetName = "Investor DTI", dti = "> 50%", grade = "B-", ltv = ws.cell(row=startRow-1, column=col).value, adj = ws.cell(row=row, column=col).value)
				else:
					adj = DTIAdj.objects.get_or_create(rateSheetName = "Investor DTI", dti = ws.cell(row=row, column=startCol-2).value, ltv = ws.cell(row=startRow-1, column=col).value, adj = ws.cell(row=row, column=col).value)
	for col in range(startCol, endCol+1):
		adj = DTIAdj.objects.get_or_create(rateSheetName = "Investor DTI", dti = "<= 36%", grade = "A", ltv = ws.cell(row=startRow-1, column=col).value, adj = 0.0)
		adj = DTIAdj.objects.get_or_create(rateSheetName = "Investor DTI", dti = "<= 36%", grade = "A-", ltv = ws.cell(row=startRow-1, column=col).value, adj = 0.0)
		adj = DTIAdj.objects.get_or_create(rateSheetName = "Investor DTI", dti = "<= 36%", grade = "B+", ltv = ws.cell(row=startRow-1, column=col).value, adj = 0.0)
		adj = DTIAdj.objects.get_or_create(rateSheetName = "Investor DTI", dti = "<= 36%", grade = "B", ltv = ws.cell(row=startRow-1, column=col).value, adj = 0.0)
		adj = DTIAdj.objects.get_or_create(rateSheetName = "Investor DTI", dti = "<= 36%", grade = "B-", ltv = ws.cell(row=startRow-1, column=col).value, adj = 0.0)
		adj = DTIAdj.objects.get_or_create(rateSheetName = "Investor DTI", dti = "36.01% - 43%", grade = "A", ltv = ws.cell(row=startRow-1, column=col).value, adj = 0.0)
		adj = DTIAdj.objects.get_or_create(rateSheetName = "Investor DTI", dti = "36.01% - 43%", grade = "A-", ltv = ws.cell(row=startRow-1, column=col).value, adj = 0.0)
		adj = DTIAdj.objects.get_or_create(rateSheetName = "Investor DTI", dti = "36.01% - 43%", grade = "B+", ltv = ws.cell(row=startRow-1, column=col).value, adj = 0.0)
		adj = DTIAdj.objects.get_or_create(rateSheetName = "Investor DTI", dti = "36.01% - 43%", grade = "B", ltv = ws.cell(row=startRow-1, column=col).value, adj = 0.0)
		adj = DTIAdj.objects.get_or_create(rateSheetName = "Investor DTI", dti = "36.01% - 43%", grade = "B-", ltv = ws.cell(row=startRow-1, column=col).value, adj = 0.0)

	#Load Loan Balance Adjustments
	startRow = 37
	endRow = 39
	startCol = 3
	endCol = 8
	for row in range(startRow, endRow+1):
		for col in range(startCol, endCol+1):
			if ws.cell(row=row, column=col).value != "NA":
				if ws.cell(row=row, column=startCol-2).value == "750k to $1mm":
					adj = LoanBalanceAdj.objects.get_or_create(rateSheetName = "Investor DTI", balance = "$750,000 - $799,999", ltv = ws.cell(row=startRow-1, column=col).value, adj = ws.cell(row=row, column=col).value)
					adj = LoanBalanceAdj.objects.get_or_create(rateSheetName = "Investor DTI", balance = "$800,000 - $999,999", ltv = ws.cell(row=startRow-1, column=col).value, adj = ws.cell(row=row, column=col).value)
				elif ws.cell(row=row, column=startCol-2).value == "<100k":
					adj = LoanBalanceAdj.objects.get_or_create(rateSheetName = "Investor DTI", balance = "< $100,000", ltv = ws.cell(row=startRow-1, column=col).value, adj = ws.cell(row=row, column=col).value)
				elif ws.cell(row=row, column=startCol-2).value == "$1mm to $1.5mm":
					adj = LoanBalanceAdj.objects.get_or_create(rateSheetName = "Investor DTI", balance = "$1,000,000 - $1,249,999", ltv = ws.cell(row=startRow-1, column=col).value, adj = ws.cell(row=row, column=col).value)
					adj = LoanBalanceAdj.objects.get_or_create(rateSheetName = "Investor DTI", balance = "$1,250,000 - $1,499,999", ltv = ws.cell(row=startRow-1, column=col).value, adj = ws.cell(row=row, column=col).value)
				else:
					adj = LoanBalanceAdj.objects.get_or_create(rateSheetName = "Investor DTI", balance = ws.cell(row=row, column=startCol-2).value, ltv = ws.cell(row=startRow-1, column=col).value, adj = ws.cell(row=row, column=col).value)
	for col in range(startCol, endCol+1):
		adj = LoanBalanceAdj.objects.get_or_create(rateSheetName = "Investor DTI", balance = "$100,000 - $149,999", ltv = ws.cell(row=startRow-1, column=col).value, adj = 0.0)
		adj = LoanBalanceAdj.objects.get_or_create(rateSheetName = "Investor DTI", balance = "$150,000 - $399,999", ltv = ws.cell(row=startRow-1, column=col).value, adj = 0.0)
		adj = LoanBalanceAdj.objects.get_or_create(rateSheetName = "Investor DTI", balance = "$400,000 - $499,999", ltv = ws.cell(row=startRow-1, column=col).value, adj = 0.0)
		adj = LoanBalanceAdj.objects.get_or_create(rateSheetName = "Investor DTI", balance = "$500,000 - $749,999", ltv = ws.cell(row=startRow-1, column=col).value, adj = 0.0)

	#Load Purpose Adjustments
	startRow = 41
	endRow = 42
	startCol = 3
	endCol = 8
	for row in range(startRow, endRow+1):
		for col in range(startCol, endCol+1):
			if ws.cell(row=row, column=col).value != "NA":
				if ws.cell(row=row, column=startCol-2).value == "Cash Out: A, A- Grades":
					adj = PurposeAdj.objects.get_or_create(rateSheetName = "Investor DTI", purpose = "Cash Out Refinance", grade = "A", ltv = ws.cell(row=startRow-1, column=col).value, adj = ws.cell(row=row, column=col).value)
					adj = PurposeAdj.objects.get_or_create(rateSheetName = "Investor DTI", purpose = "Cash Out Refinance", grade = "A-", ltv = ws.cell(row=startRow-1, column=col).value, adj = ws.cell(row=row, column=col).value)
				elif ws.cell(row=row, column=startCol-2).value == "Cash Out: B+, B, B- Grades And Foreign Credit":
					adj = PurposeAdj.objects.get_or_create(rateSheetName = "Investor DTI", purpose = "Cash Out Refinance", grade = "B+", ltv = ws.cell(row=startRow-1, column=col).value, adj = ws.cell(row=row, column=col).value)			
					adj = PurposeAdj.objects.get_or_create(rateSheetName = "Investor DTI", purpose = "Cash Out Refinance", grade = "B", ltv = ws.cell(row=startRow-1, column=col).value, adj = ws.cell(row=row, column=col).value)
					adj = PurposeAdj.objects.get_or_create(rateSheetName = "Investor DTI", purpose = "Cash Out Refinance", grade = "B-", ltv = ws.cell(row=startRow-1, column=col).value, adj = ws.cell(row=row, column=col).value)
	for col in range(startCol, endCol+1):
		adj = PurposeAdj.objects.get_or_create(rateSheetName = "Investor DTI", purpose = "Purchase", grade = "A", ltv = ws.cell(row=startRow-1, column=col).value, adj = 0.0)
		adj = PurposeAdj.objects.get_or_create(rateSheetName = "Investor DTI", purpose = "Purchase", grade = "A-", ltv = ws.cell(row=startRow-1, column=col).value, adj = 0.0)
		adj = PurposeAdj.objects.get_or_create(rateSheetName = "Investor DTI", purpose = "Purchase", grade = "B+", ltv = ws.cell(row=startRow-1, column=col).value, adj = 0.0)
		adj = PurposeAdj.objects.get_or_create(rateSheetName = "Investor DTI", purpose = "Purchase", grade = "B", ltv = ws.cell(row=startRow-1, column=col).value, adj = 0.0)
		adj = PurposeAdj.objects.get_or_create(rateSheetName = "Investor DTI", purpose = "Purchase", grade = "B-", ltv = ws.cell(row=startRow-1, column=col).value, adj = 0.0)
		adj = PurposeAdj.objects.get_or_create(rateSheetName = "Investor DTI", purpose = "Rate/Term Refinance", grade = "A", ltv = ws.cell(row=startRow-1, column=col).value, adj = 0.0)
		adj = PurposeAdj.objects.get_or_create(rateSheetName = "Investor DTI", purpose = "Rate/Term Refinance", grade = "A-", ltv = ws.cell(row=startRow-1, column=col).value, adj = 0.0)
		adj = PurposeAdj.objects.get_or_create(rateSheetName = "Investor DTI", purpose = "Rate/Term Refinance", grade = "B+", ltv = ws.cell(row=startRow-1, column=col).value, adj = 0.0)
		adj = PurposeAdj.objects.get_or_create(rateSheetName = "Investor DTI", purpose = "Rate/Term Refinance", grade = "B", ltv = ws.cell(row=startRow-1, column=col).value, adj = 0.0)
		adj = PurposeAdj.objects.get_or_create(rateSheetName = "Investor DTI", purpose = "Rate/Term Refinance", grade = "B-", ltv = ws.cell(row=startRow-1, column=col).value, adj = 0.0)


	#Load Property Type Adjustments
	startRow = 44
	endRow = 46
	startCol = 3
	endCol = 8
	for row in range(startRow, endRow+1):
		for col in range(startCol, endCol+1):
			if ws.cell(row=row, column=col).value != "NA":
				adj = PropertyTypeAdj.objects.get_or_create(rateSheetName = "Investor DTI", propertyType = ws.cell(row=row, column=startCol-2).value, ltv = ws.cell(row=startRow-1, column=col).value, adj = ws.cell(row=row, column=col).value)
	for col in range(startCol, endCol+1):
		adj = PropertyTypeAdj.objects.get_or_create(rateSheetName = "Investor DTI", propertyType = "SFR", ltv = ws.cell(row=startRow-1, column=col).value, adj = 0.0)
		adj = PropertyTypeAdj.objects.get_or_create(rateSheetName = "Investor DTI", propertyType = "PUD", ltv = ws.cell(row=startRow-1, column=col).value, adj = 0.0)

	#Load State Adjustments
	startRow = 48
	endRow = 50
	startCol = 3
	endCol = 8
	for row in range(startRow, endRow+1):
		for col in range(startCol, endCol+1):
			if ws.cell(row=row, column=col).value != "NA":
				adj = StateAdj.objects.get_or_create(rateSheetName = "Investor DTI", state = ws.cell(row=row, column=startCol-2).value, ltv = ws.cell(row=startRow-1, column=col).value, adj = ws.cell(row=row, column=col).value)

	#Load Interest Only Adjustments
	startRow = 52
	endRow = 52
	startCol = 3
	endCol = 8
	for row in range(startRow, endRow+1):
		for col in range(startCol, endCol+1):
			if ws.cell(row=row, column=col).value != "NA":
				adj = IOAdj.objects.get_or_create(rateSheetName = "Investor DTI", io = "Yes", ltv = ws.cell(row=startRow-1, column=col).value, adj = ws.cell(row=row, column=col).value)
	for col in range(startCol, endCol+1):
		adj = IOAdj.objects.get_or_create(rateSheetName = "Investor DTI", io = "No", ltv = ws.cell(row=startRow-1, column=col).value, adj = 0.0)

	#Load Prepayment Adjustments
	startRow = 54
	endRow = 57
	startCol = 3
	endCol = 8
	for row in range(startRow, endRow+1):
		for col in range(startCol, endCol+1):
			if ws.cell(row=row, column=col).value != "NA":
				adj = PrepayAdj.objects.get_or_create(rateSheetName = "Investor DTI", prepayTerm = ws.cell(row=row, column=startCol-2).value, ltv = ws.cell(row=startRow-1, column=col).value, adj = ws.cell(row=row, column=col).value)
	for col in range(startCol, endCol+1):
		adj = PrepayAdj.objects.get_or_create(rateSheetName = "Investor DTI", prepayTerm = "N/A", ltv = ws.cell(row=startRow-1, column=col).value, adj = 0.0)

	#Load Loan Term Adjustments
	adj = LoanTermAdj.objects.get_or_create(rateSheetName = "Investor DTI", loanTerm = "5/1 ARM" , adj = 0.0)
	adj = LoanTermAdj.objects.get_or_create(rateSheetName = "Investor DTI", loanTerm = "7/1 ARM" , adj = 0.1)
	adj = LoanTermAdj.objects.get_or_create(rateSheetName = "Investor DTI", loanTerm = "15 Yr Fix" , adj = 0.1)
	adj = LoanTermAdj.objects.get_or_create(rateSheetName = "Investor DTI", loanTerm = "30 Yr Fix" , adj = 0.3)

	# INVESTOR PI RATE SHEET
	print "Loading Verus Investor PI Rate Sheet..."
	wb = load_workbook(filename = '/home/invictus/VerusPricer/static/rates/Verus Investor PI Rate Sheet.xlsx')
	ws = wb.active
	
	#Load Base Rates
	startRow = 9
	endRow = 16
	startCol = 2
	endCol = 7
	for row in range(startRow, endRow+1):
		for col in range(startCol, endCol+1):
			if ws.cell(row=row, column=col).value:
				rate = BaseRate.objects.get_or_create(rateSheetName = "Investor PI", fico = ws.cell(row=row, column=startCol-1).value, ltv = ws.cell(row=startRow-1, column=col).value, adj = ws.cell(row=row, column=col).value)
				#rate.save()

	#Load DocType Adjustments
	startRow = 20
	endRow = 21
	startCol = 2
	endCol = 7
	for row in range(startRow, endRow+1):
		for col in range(startCol, endCol+1):
			if ws.cell(row=row, column=col).value != "NA":
				if ws.cell(row=row, column=startCol-1).value == "DSCR (620 Min FICO)":
					adj = DocTypeAdj.objects.get_or_create(rateSheetName = "Investor PI", docType = "DSCR", grade = "A", ltv = ws.cell(row=startRow-1, column=col).value, adj = ws.cell(row=row, column=col).value)
				elif ws.cell(row=row, column=startCol-1).value == "No Ratio (640 Min FICO)":
					adj = DocTypeAdj.objects.get_or_create(rateSheetName = "Investor PI", docType = "No Ratio", grade = "A", ltv = ws.cell(row=startRow-1, column=col).value, adj = ws.cell(row=row, column=col).value)
				else:
					adj = DocTypeAdj.objects.get_or_create(rateSheetName = "Investor PI", docType = ws.cell(row=row, column=startCol-1).value, grade = "A", ltv = ws.cell(row=startRow-1, column=col).value, adj = ws.cell(row=row, column=col).value)

	#Load Reserve Adjustments
	startRow = 23
	endRow = 23
	startCol = 2
	endCol = 7
	for row in range(startRow, endRow+1):
		for col in range(startCol, endCol+1):
			if ws.cell(row=row, column=col).value != "NA":
				if ws.cell(row=row, column=startCol-1).value == "< 6 Months":
					adj = ReservesAdj.objects.get_or_create(rateSheetName = "Investor PI", reserves = ws.cell(row=row, column=startCol-1).value, ltv = ws.cell(row=startRow-1, column=col).value, adj = ws.cell(row=row, column=col).value)
					adj = ReservesAdj.objects.get_or_create(rateSheetName = "Investor PI", reserves = "None", ltv = ws.cell(row=startRow-1, column=col).value, adj = ws.cell(row=row, column=col).value)

	for col in range(startCol, endCol+1):
		adj = ReservesAdj.objects.get_or_create(rateSheetName = "Investor PI", reserves = "6 - 11 Months", ltv = ws.cell(row=startRow-1, column=col).value, adj = 0.0)
		adj = ReservesAdj.objects.get_or_create(rateSheetName = "Investor PI", reserves = "12 - 17 Months", ltv = ws.cell(row=startRow-1, column=col).value, adj = 0.0)
		adj = ReservesAdj.objects.get_or_create(rateSheetName = "Investor PI", reserves = "18 - 23 Months", ltv = ws.cell(row=startRow-1, column=col).value, adj = 0.0)
		adj = ReservesAdj.objects.get_or_create(rateSheetName = "Investor PI", reserves = ">= 24 Months", ltv = ws.cell(row=startRow-1, column=col).value, adj = 0.0)
		adj = ReservesAdj.objects.get_or_create(rateSheetName = "Investor PI", reserves = "N/A", ltv = ws.cell(row=startRow-1, column=col).value, adj = 0.0)

	#Load Loan Balance Adjustments
	startRow = 25
	endRow = 28
	startCol = 2
	endCol = 7
	for row in range(startRow, endRow+1):
		for col in range(startCol, endCol+1):
			if ws.cell(row=row, column=col).value != "NA":
				if ws.cell(row=row, column=startCol-1).value == "750k to $1mm":
					adj = LoanBalanceAdj.objects.get_or_create(rateSheetName = "Investor PI", balance = "$750,000 - $799,999", ltv = ws.cell(row=startRow-1, column=col).value, adj = ws.cell(row=row, column=col).value)
					adj = LoanBalanceAdj.objects.get_or_create(rateSheetName = "Investor PI", balance = "$800,000 - $999,999", ltv = ws.cell(row=startRow-1, column=col).value, adj = ws.cell(row=row, column=col).value)
				elif ws.cell(row=row, column=startCol-1).value == "<100k":
					adj = LoanBalanceAdj.objects.get_or_create(rateSheetName = "Investor PI", balance = "< $100,000", ltv = ws.cell(row=startRow-1, column=col).value, adj = ws.cell(row=row, column=col).value)
				elif ws.cell(row=row, column=startCol-1).value == "$1mm to $1.5mm":
					adj = LoanBalanceAdj.objects.get_or_create(rateSheetName = "Investor PI", balance = "$1,000,000 - $1,249,999", ltv = ws.cell(row=startRow-1, column=col).value, adj = ws.cell(row=row, column=col).value)
					adj = LoanBalanceAdj.objects.get_or_create(rateSheetName = "Investor PI", balance = "$1,250,000 - $1,499,999", ltv = ws.cell(row=startRow-1, column=col).value, adj = ws.cell(row=row, column=col).value)
				elif ws.cell(row=row, column=startCol-1).value == ">= $1.5mm":
					adj = LoanBalanceAdj.objects.get_or_create(rateSheetName = "Investor PI", balance = "$1,500,000 - $2,000,000", ltv = ws.cell(row=startRow-1, column=col).value, adj = ws.cell(row=row, column=col).value)
				else:
					adj = LoanBalanceAdj.objects.get_or_create(rateSheetName = "Investor PI", balance = ws.cell(row=row, column=startCol-1).value, ltv = ws.cell(row=startRow-1, column=col).value, adj = ws.cell(row=row, column=col).value)
	for col in range(startCol, endCol+1):
		adj = LoanBalanceAdj.objects.get_or_create(rateSheetName = "Investor PI", balance = "$100,000 - $149,999", ltv = ws.cell(row=startRow-1, column=col).value, adj = 0.0)
		adj = LoanBalanceAdj.objects.get_or_create(rateSheetName = "Investor PI", balance = "$150,000 - $399,999", ltv = ws.cell(row=startRow-1, column=col).value, adj = 0.0)
		adj = LoanBalanceAdj.objects.get_or_create(rateSheetName = "Investor PI", balance = "$400,000 - $499,999", ltv = ws.cell(row=startRow-1, column=col).value, adj = 0.0)
		adj = LoanBalanceAdj.objects.get_or_create(rateSheetName = "Investor PI", balance = "$500,000 - $749,999", ltv = ws.cell(row=startRow-1, column=col).value, adj = 0.0)

	#Load Purpose Adjustments
	startRow = 30
	endRow = 30
	startCol = 2
	endCol = 7
	for row in range(startRow, endRow+1):
		for col in range(startCol, endCol+1):
			if ws.cell(row=row, column=col).value != "NA":
				if ws.cell(row=row, column=startCol-1).value == "Cash Out":
					adj = PurposeAdj.objects.get_or_create(rateSheetName = "Investor PI", purpose = "Cash Out Refinance", grade = "A", ltv = ws.cell(row=startRow-1, column=col).value, adj = ws.cell(row=row, column=col).value)
	for col in range(startCol, endCol+1):
		adj = PurposeAdj.objects.get_or_create(rateSheetName = "Investor PI", purpose = "Purchase", grade = "A", ltv = ws.cell(row=startRow-1, column=col).value, adj = 0.0)
		adj = PurposeAdj.objects.get_or_create(rateSheetName = "Investor PI", purpose = "Rate/Term Refinance", grade = "A", ltv = ws.cell(row=startRow-1, column=col).value, adj = 0.0)


	#Load Property Type Adjustments
	startRow = 32
	endRow = 33
	startCol = 2
	endCol = 7
	for row in range(startRow, endRow+1):
		for col in range(startCol, endCol+1):
			if ws.cell(row=row, column=col).value != "NA":
				adj = PropertyTypeAdj.objects.get_or_create(rateSheetName = "Investor PI", propertyType = ws.cell(row=row, column=startCol-1).value, ltv = ws.cell(row=startRow-1, column=col).value, adj = ws.cell(row=row, column=col).value)
	for col in range(startCol, endCol+1):
		adj = PropertyTypeAdj.objects.get_or_create(rateSheetName = "Investor PI", propertyType = "SFR", ltv = ws.cell(row=startRow-1, column=col).value, adj = 0.0)
		adj = PropertyTypeAdj.objects.get_or_create(rateSheetName = "Investor PI", propertyType = "PUD", ltv = ws.cell(row=startRow-1, column=col).value, adj = 0.0)

	#Load State Adjustments
	startRow = 35
	endRow = 37
	startCol = 2
	endCol = 7
	for row in range(startRow, endRow+1):
		for col in range(startCol, endCol+1):
			if ws.cell(row=row, column=col).value != "NA":
				adj = StateAdj.objects.get_or_create(rateSheetName = "Investor PI", state = ws.cell(row=row, column=startCol-1).value, ltv = ws.cell(row=startRow-1, column=col).value, adj = ws.cell(row=row, column=col).value)

	#Load Interest Only Adjustments
	startRow = 39
	endRow = 39
	startCol = 2
	endCol = 7
	for row in range(startRow, endRow+1):
		for col in range(startCol, endCol+1):
			if ws.cell(row=row, column=col).value != "NA":
				adj = IOAdj.objects.get_or_create(rateSheetName = "Investor PI", io = "Yes", ltv = ws.cell(row=startRow-1, column=col).value, adj = ws.cell(row=row, column=col).value)
	for col in range(startCol, endCol+1):
		adj = IOAdj.objects.get_or_create(rateSheetName = "Investor PI", io = "No", ltv = ws.cell(row=startRow-1, column=col).value, adj = 0.0)

	#Load Prepayment Adjustments
	startRow = 41
	endRow = 44
	startCol = 2
	endCol = 7
	for row in range(startRow, endRow+1):
		for col in range(startCol, endCol+1):
			if ws.cell(row=row, column=col).value != "NA":
				adj = PrepayAdj.objects.get_or_create(rateSheetName = "Investor PI", prepayTerm = ws.cell(row=row, column=startCol-1).value, ltv = ws.cell(row=startRow-1, column=col).value, adj = ws.cell(row=row, column=col).value)
	for col in range(startCol, endCol+1):
		adj = PrepayAdj.objects.get_or_create(rateSheetName = "Investor PI", prepayTerm = "N/A", ltv = ws.cell(row=startRow-1, column=col).value, adj = 0.0)

	#Load Loan Term Adjustments
	adj = LoanTermAdj.objects.get_or_create(rateSheetName = "Investor PI", loanTerm = "3/1 ARM" , adj = 0.0)
	adj = LoanTermAdj.objects.get_or_create(rateSheetName = "Investor PI", loanTerm = "5/1 ARM" , adj = 0.1)
	adj = LoanTermAdj.objects.get_or_create(rateSheetName = "Investor PI", loanTerm = "7/1 ARM" , adj = 0.2)
	adj = LoanTermAdj.objects.get_or_create(rateSheetName = "Investor PI", loanTerm = "15 Yr Fix" , adj = 0.2)
	adj = LoanTermAdj.objects.get_or_create(rateSheetName = "Investor PI", loanTerm = "30 Yr Fix" , adj = 0.4)

if __name__ == '__main__':
	print "Importing Rate Sheets..."
	importRateSheets()