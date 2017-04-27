import os, sys
os.environ.setdefault('DJANGO_SETTINGS_MODULE', 'VerusPricer.settings')
import django
django.setup()
from pricer.models import CAProductMatrix, DTIProductMatrix, PIProductMatrix, FNProductMatrix
from datetime import datetime
from openpyxl import *

# Helper function to map matrix loan amounts to form/rate sheet loan amounts
def createCAModels(occupancy, grade, fico, loanamount, docType, purpose, maxLTV):
	if loanamount == "<=500,000":
		result = CAProductMatrix.objects.get_or_create(occupancy = occupancy, grade = grade, fico = fico, loanamount = "< $100,000", docType = docType, purpose = purpose, maxLTV = maxLTV)
		result = CAProductMatrix.objects.get_or_create(occupancy = occupancy, grade = grade, fico = fico, loanamount = "$100,000 - $149,999", docType = docType, purpose = purpose, maxLTV = maxLTV)
		result = CAProductMatrix.objects.get_or_create(occupancy = occupancy, grade = grade, fico = fico, loanamount = "$150,000 - $399,999", docType = docType, purpose = purpose, maxLTV = maxLTV)
		result = CAProductMatrix.objects.get_or_create(occupancy = occupancy, grade = grade, fico = fico, loanamount = "$400,000 - $499,999", docType = docType, purpose = purpose, maxLTV = maxLTV)
	elif loanamount == "<=800,000":
		result = CAProductMatrix.objects.get_or_create(occupancy = occupancy, grade = grade, fico = fico, loanamount = "< $100,000", docType = docType, purpose = purpose, maxLTV = maxLTV)
		result = CAProductMatrix.objects.get_or_create(occupancy = occupancy, grade = grade, fico = fico, loanamount = "$100,000 - $149,999", docType = docType, purpose = purpose, maxLTV = maxLTV)
		result = CAProductMatrix.objects.get_or_create(occupancy = occupancy, grade = grade, fico = fico, loanamount = "$150,000 - $399,999", docType = docType, purpose = purpose, maxLTV = maxLTV)
		result = CAProductMatrix.objects.get_or_create(occupancy = occupancy, grade = grade, fico = fico, loanamount = "$400,000 - $499,999", docType = docType, purpose = purpose, maxLTV = maxLTV)
		result = CAProductMatrix.objects.get_or_create(occupancy = occupancy, grade = grade, fico = fico, loanamount = "$500,000 - $749,999", docType = docType, purpose = purpose, maxLTV = maxLTV)
		result = CAProductMatrix.objects.get_or_create(occupancy = occupancy, grade = grade, fico = fico, loanamount = "$750,000 - $799,999", docType = docType, purpose = purpose, maxLTV = maxLTV)
	elif loanamount == "<=1,000,000":
		result = CAProductMatrix.objects.get_or_create(occupancy = occupancy, grade = grade, fico = fico, loanamount = "< $100,000", docType = docType, purpose = purpose, maxLTV = maxLTV)
		result = CAProductMatrix.objects.get_or_create(occupancy = occupancy, grade = grade, fico = fico, loanamount = "$100,000 - $149,999", docType = docType, purpose = purpose, maxLTV = maxLTV)
		result = CAProductMatrix.objects.get_or_create(occupancy = occupancy, grade = grade, fico = fico, loanamount = "$150,000 - $399,999", docType = docType, purpose = purpose, maxLTV = maxLTV)
		result = CAProductMatrix.objects.get_or_create(occupancy = occupancy, grade = grade, fico = fico, loanamount = "$400,000 - $499,999", docType = docType, purpose = purpose, maxLTV = maxLTV)
		result = CAProductMatrix.objects.get_or_create(occupancy = occupancy, grade = grade, fico = fico, loanamount = "$500,000 - $749,999", docType = docType, purpose = purpose, maxLTV = maxLTV)
		result = CAProductMatrix.objects.get_or_create(occupancy = occupancy, grade = grade, fico = fico, loanamount = "$750,000 - $799,999", docType = docType, purpose = purpose, maxLTV = maxLTV)
		result = CAProductMatrix.objects.get_or_create(occupancy = occupancy, grade = grade, fico = fico, loanamount = "$800,000 - $999,999", docType = docType, purpose = purpose, maxLTV = maxLTV)
	elif loanamount == "<=1,250,000":
		result = CAProductMatrix.objects.get_or_create(occupancy = occupancy, grade = grade, fico = fico, loanamount = "< $100,000", docType = docType, purpose = purpose, maxLTV = maxLTV)
		result = CAProductMatrix.objects.get_or_create(occupancy = occupancy, grade = grade, fico = fico, loanamount = "$100,000 - $149,999", docType = docType, purpose = purpose, maxLTV = maxLTV)
		result = CAProductMatrix.objects.get_or_create(occupancy = occupancy, grade = grade, fico = fico, loanamount = "$150,000 - $399,999", docType = docType, purpose = purpose, maxLTV = maxLTV)
		result = CAProductMatrix.objects.get_or_create(occupancy = occupancy, grade = grade, fico = fico, loanamount = "$400,000 - $499,999", docType = docType, purpose = purpose, maxLTV = maxLTV)
		result = CAProductMatrix.objects.get_or_create(occupancy = occupancy, grade = grade, fico = fico, loanamount = "$500,000 - $749,999", docType = docType, purpose = purpose, maxLTV = maxLTV)
		result = CAProductMatrix.objects.get_or_create(occupancy = occupancy, grade = grade, fico = fico, loanamount = "$750,000 - $799,999", docType = docType, purpose = purpose, maxLTV = maxLTV)
		result = CAProductMatrix.objects.get_or_create(occupancy = occupancy, grade = grade, fico = fico, loanamount = "$800,000 - $999,999", docType = docType, purpose = purpose, maxLTV = maxLTV)
		result = CAProductMatrix.objects.get_or_create(occupancy = occupancy, grade = grade, fico = fico, loanamount = "$1,000,000 - $1,249,999", docType = docType, purpose = purpose, maxLTV = maxLTV)
	elif loanamount == "1,000,001 - 1,500,000":
		result = CAProductMatrix.objects.get_or_create(occupancy = occupancy, grade = grade, fico = fico, loanamount = "$1,000,000 - $1,249,999", docType = docType, purpose = purpose, maxLTV = maxLTV)
		result = CAProductMatrix.objects.get_or_create(occupancy = occupancy, grade = grade, fico = fico, loanamount = "$1,250,000 - $1,499,999", docType = docType, purpose = purpose, maxLTV = maxLTV)
	elif loanamount == "1,500,001 - 2,000,000":
		result = CAProductMatrix.objects.get_or_create(occupancy = occupancy, grade = grade, fico = fico, loanamount = "$1,500,000 - $2,000,000", docType = docType, purpose = purpose, maxLTV = maxLTV)

# Helper function to map matrix loan amounts to form/rate sheet loan amounts
def createDTIModels(grade, fico, loanamount, docType, purpose, maxLTV):
	if loanamount == "<=500,000":
		result = DTIProductMatrix.objects.get_or_create(grade = grade, fico = fico, loanamount = "< $100,000", docType = docType, purpose = purpose, maxLTV = maxLTV)
		result = DTIProductMatrix.objects.get_or_create(grade = grade, fico = fico, loanamount = "$100,000 - $149,999", docType = docType, purpose = purpose, maxLTV = maxLTV)
		result = DTIProductMatrix.objects.get_or_create(grade = grade, fico = fico, loanamount = "$150,000 - $399,999", docType = docType, purpose = purpose, maxLTV = maxLTV)
		result = DTIProductMatrix.objects.get_or_create(grade = grade, fico = fico, loanamount = "$400,000 - $499,999", docType = docType, purpose = purpose, maxLTV = maxLTV)
	elif loanamount == "<=800,000":
		result = DTIProductMatrix.objects.get_or_create(grade = grade, fico = fico, loanamount = "< $100,000", docType = docType, purpose = purpose, maxLTV = maxLTV)
		result = DTIProductMatrix.objects.get_or_create(grade = grade, fico = fico, loanamount = "$100,000 - $149,999", docType = docType, purpose = purpose, maxLTV = maxLTV)
		result = DTIProductMatrix.objects.get_or_create(grade = grade, fico = fico, loanamount = "$150,000 - $399,999", docType = docType, purpose = purpose, maxLTV = maxLTV)
		result = DTIProductMatrix.objects.get_or_create(grade = grade, fico = fico, loanamount = "$400,000 - $499,999", docType = docType, purpose = purpose, maxLTV = maxLTV)
		result = DTIProductMatrix.objects.get_or_create(grade = grade, fico = fico, loanamount = "$500,000 - $749,999", docType = docType, purpose = purpose, maxLTV = maxLTV)
		result = DTIProductMatrix.objects.get_or_create(grade = grade, fico = fico, loanamount = "$750,000 - $799,999", docType = docType, purpose = purpose, maxLTV = maxLTV)
	elif loanamount == "<=1,000,000":
		result = DTIProductMatrix.objects.get_or_create(grade = grade, fico = fico, loanamount = "< $100,000", docType = docType, purpose = purpose, maxLTV = maxLTV)
		result = DTIProductMatrix.objects.get_or_create(grade = grade, fico = fico, loanamount = "$100,000 - $149,999", docType = docType, purpose = purpose, maxLTV = maxLTV)
		result = DTIProductMatrix.objects.get_or_create(grade = grade, fico = fico, loanamount = "$150,000 - $399,999", docType = docType, purpose = purpose, maxLTV = maxLTV)
		result = DTIProductMatrix.objects.get_or_create(grade = grade, fico = fico, loanamount = "$400,000 - $499,999", docType = docType, purpose = purpose, maxLTV = maxLTV)
		result = DTIProductMatrix.objects.get_or_create(grade = grade, fico = fico, loanamount = "$500,000 - $749,999", docType = docType, purpose = purpose, maxLTV = maxLTV)
		result = DTIProductMatrix.objects.get_or_create(grade = grade, fico = fico, loanamount = "$750,000 - $799,999", docType = docType, purpose = purpose, maxLTV = maxLTV)
		result = DTIProductMatrix.objects.get_or_create(grade = grade, fico = fico, loanamount = "$800,000 - $999,999", docType = docType, purpose = purpose, maxLTV = maxLTV)
	elif loanamount == "<=1,250,000":
		result = DTIProductMatrix.objects.get_or_create(grade = grade, fico = fico, loanamount = "< $100,000", docType = docType, purpose = purpose, maxLTV = maxLTV)
		result = DTIProductMatrix.objects.get_or_create(grade = grade, fico = fico, loanamount = "$100,000 - $149,999", docType = docType, purpose = purpose, maxLTV = maxLTV)
		result = DTIProductMatrix.objects.get_or_create(grade = grade, fico = fico, loanamount = "$150,000 - $399,999", docType = docType, purpose = purpose, maxLTV = maxLTV)
		result = DTIProductMatrix.objects.get_or_create(grade = grade, fico = fico, loanamount = "$400,000 - $499,999", docType = docType, purpose = purpose, maxLTV = maxLTV)
		result = DTIProductMatrix.objects.get_or_create(grade = grade, fico = fico, loanamount = "$500,000 - $749,999", docType = docType, purpose = purpose, maxLTV = maxLTV)
		result = DTIProductMatrix.objects.get_or_create(grade = grade, fico = fico, loanamount = "$750,000 - $799,999", docType = docType, purpose = purpose, maxLTV = maxLTV)
		result = DTIProductMatrix.objects.get_or_create(grade = grade, fico = fico, loanamount = "$800,000 - $999,999", docType = docType, purpose = purpose, maxLTV = maxLTV)
		result = DTIProductMatrix.objects.get_or_create(grade = grade, fico = fico, loanamount = "$1,000,000 - $1,249,999", docType = docType, purpose = purpose, maxLTV = maxLTV)
	elif loanamount == "1,000,001 - 1,500,000":
		result = DTIProductMatrix.objects.get_or_create(grade = grade, fico = fico, loanamount = "$1,000,000 - $1,249,999", docType = docType, purpose = purpose, maxLTV = maxLTV)
		result = DTIProductMatrix.objects.get_or_create(grade = grade, fico = fico, loanamount = "$1,250,000 - $1,499,999", docType = docType, purpose = purpose, maxLTV = maxLTV)
	elif loanamount == "1,500,001 - 2,000,000":
		result = DTIProductMatrix.objects.get_or_create(grade = grade, fico = fico, loanamount = "$1,500,000 - $2,000,000", docType = docType, purpose = purpose, maxLTV = maxLTV)

# Helper function to map matrix loan amounts to form/rate sheet loan amounts
def createPIModels(docType, fico, loanamount, reserves, purpose, maxLTV):
	if loanamount == "<=500,000":
		result = PIProductMatrix.objects.get_or_create(docType = docType, fico = fico, loanamount = "< $100,000", reserves = reserves, purpose = purpose, maxLTV = maxLTV)
		result = PIProductMatrix.objects.get_or_create(docType = docType, fico = fico, loanamount = "$100,000 - $149,999", reserves = reserves, purpose = purpose, maxLTV = maxLTV)
		result = PIProductMatrix.objects.get_or_create(docType = docType, fico = fico, loanamount = "$150,000 - $399,999", reserves = reserves, purpose = purpose, maxLTV = maxLTV)
		result = PIProductMatrix.objects.get_or_create(docType = docType, fico = fico, loanamount = "$400,000 - $499,999", reserves = reserves, purpose = purpose, maxLTV = maxLTV)
	elif loanamount == "<=800,000":
		result = PIProductMatrix.objects.get_or_create(docType = docType, fico = fico, loanamount = "< $100,000", reserves = reserves, purpose = purpose, maxLTV = maxLTV)
		result = PIProductMatrix.objects.get_or_create(docType = docType, fico = fico, loanamount = "$100,000 - $149,999", reserves = reserves, purpose = purpose, maxLTV = maxLTV)
		result = PIProductMatrix.objects.get_or_create(docType = docType, fico = fico, loanamount = "$150,000 - $399,999", reserves = reserves, purpose = purpose, maxLTV = maxLTV)
		result = PIProductMatrix.objects.get_or_create(docType = docType, fico = fico, loanamount = "$400,000 - $499,999", reserves = reserves, purpose = purpose, maxLTV = maxLTV)
		result = PIProductMatrix.objects.get_or_create(docType = docType, fico = fico, loanamount = "$500,000 - $749,999", reserves = reserves, purpose = purpose, maxLTV = maxLTV)
		result = PIProductMatrix.objects.get_or_create(docType = docType, fico = fico, loanamount = "$750,000 - $799,999", reserves = reserves, purpose = purpose, maxLTV = maxLTV)
	elif loanamount == "<=1,000,000":
		result = PIProductMatrix.objects.get_or_create(docType = docType, fico = fico, loanamount = "< $100,000", reserves = reserves, purpose = purpose, maxLTV = maxLTV)
		result = PIProductMatrix.objects.get_or_create(docType = docType, fico = fico, loanamount = "$100,000 - $149,999", reserves = reserves, purpose = purpose, maxLTV = maxLTV)
		result = PIProductMatrix.objects.get_or_create(docType = docType, fico = fico, loanamount = "$150,000 - $399,999", reserves = reserves, purpose = purpose, maxLTV = maxLTV)
		result = PIProductMatrix.objects.get_or_create(docType = docType, fico = fico, loanamount = "$400,000 - $499,999", reserves = reserves, purpose = purpose, maxLTV = maxLTV)
		result = PIProductMatrix.objects.get_or_create(docType = docType, fico = fico, loanamount = "$500,000 - $749,999", reserves = reserves, purpose = purpose, maxLTV = maxLTV)
		result = PIProductMatrix.objects.get_or_create(docType = docType, fico = fico, loanamount = "$750,000 - $799,999", reserves = reserves, purpose = purpose, maxLTV = maxLTV)
		result = PIProductMatrix.objects.get_or_create(docType = docType, fico = fico, loanamount = "$800,000 - $999,999", reserves = reserves, purpose = purpose, maxLTV = maxLTV)
	elif loanamount == "<=1,250,000":
		result = PIProductMatrix.objects.get_or_create(docType = docType, fico = fico, loanamount = "< $100,000", reserves = reserves, purpose = purpose, maxLTV = maxLTV)
		result = PIProductMatrix.objects.get_or_create(docType = docType, fico = fico, loanamount = "$100,000 - $149,999", reserves = reserves, purpose = purpose, maxLTV = maxLTV)
		result = PIProductMatrix.objects.get_or_create(docType = docType, fico = fico, loanamount = "$150,000 - $399,999", reserves = reserves, purpose = purpose, maxLTV = maxLTV)
		result = PIProductMatrix.objects.get_or_create(docType = docType, fico = fico, loanamount = "$400,000 - $499,999", reserves = reserves, purpose = purpose, maxLTV = maxLTV)
		result = PIProductMatrix.objects.get_or_create(docType = docType, fico = fico, loanamount = "$500,000 - $749,999", reserves = reserves, purpose = purpose, maxLTV = maxLTV)
		result = PIProductMatrix.objects.get_or_create(docType = docType, fico = fico, loanamount = "$750,000 - $799,999", reserves = reserves, purpose = purpose, maxLTV = maxLTV)
		result = PIProductMatrix.objects.get_or_create(docType = docType, fico = fico, loanamount = "$800,000 - $999,999", reserves = reserves, purpose = purpose, maxLTV = maxLTV)
		result = PIProductMatrix.objects.get_or_create(docType = docType, fico = fico, loanamount = "$1,000,000 - $1,249,999", reserves = reserves, purpose = purpose, maxLTV = maxLTV)
	elif loanamount == "1,000,001 - 1,500,000":
		result = PIProductMatrix.objects.get_or_create(docType = docType, fico = fico, loanamount = "$1,000,000 - $1,249,999", reserves = reserves, purpose = purpose, maxLTV = maxLTV)
		result = PIProductMatrix.objects.get_or_create(docType = docType, fico = fico, loanamount = "$1,250,000 - $1,499,999", reserves = reserves, purpose = purpose, maxLTV = maxLTV)
	elif loanamount == "1,500,001 - 2,000,000":
		result = PIProductMatrix.objects.get_or_create(docType = docType, fico = fico, loanamount = "$1,500,000 - $2,000,000", reserves = reserves, purpose = purpose, maxLTV = maxLTV)

# Helper function to map matrix loan amounts to form/rate sheet loan amounts
def createFNModels(docType, grade, loanamount, purpose, maxLTV):
	if loanamount == "<=500,000":
		result = FNProductMatrix.objects.get_or_create(docType = docType, grade = grade, loanamount = "< $100,000", purpose = purpose, maxLTV = maxLTV)
		result = FNProductMatrix.objects.get_or_create(docType = docType, grade = grade, loanamount = "$100,000 - $149,999", purpose = purpose, maxLTV = maxLTV)
		result = FNProductMatrix.objects.get_or_create(docType = docType, grade = grade, loanamount = "$150,000 - $399,999", purpose = purpose, maxLTV = maxLTV)
		result = FNProductMatrix.objects.get_or_create(docType = docType, grade = grade, loanamount = "$400,000 - $499,999", purpose = purpose, maxLTV = maxLTV)
	elif loanamount == "<=800,000":
		result = FNProductMatrix.objects.get_or_create(docType = docType, grade = grade, loanamount = "< $100,000", purpose = purpose, maxLTV = maxLTV)
		result = FNProductMatrix.objects.get_or_create(docType = docType, grade = grade, loanamount = "$100,000 - $149,999", purpose = purpose, maxLTV = maxLTV)
		result = FNProductMatrix.objects.get_or_create(docType = docType, grade = grade, loanamount = "$150,000 - $399,999", purpose = purpose, maxLTV = maxLTV)
		result = FNProductMatrix.objects.get_or_create(docType = docType, grade = grade, loanamount = "$400,000 - $499,999", purpose = purpose, maxLTV = maxLTV)
		result = FNProductMatrix.objects.get_or_create(docType = docType, grade = grade, loanamount = "$500,000 - $749,999", purpose = purpose, maxLTV = maxLTV)
		result = FNProductMatrix.objects.get_or_create(docType = docType, grade = grade, loanamount = "$750,000 - $799,999", purpose = purpose, maxLTV = maxLTV)
	elif loanamount == "<=1,000,000":
		result = FNProductMatrix.objects.get_or_create(docType = docType, grade = grade, loanamount = "< $100,000", purpose = purpose, maxLTV = maxLTV)
		result = FNProductMatrix.objects.get_or_create(docType = docType, grade = grade, loanamount = "$100,000 - $149,999", purpose = purpose, maxLTV = maxLTV)
		result = FNProductMatrix.objects.get_or_create(docType = docType, grade = grade, loanamount = "$150,000 - $399,999", purpose = purpose, maxLTV = maxLTV)
		result = FNProductMatrix.objects.get_or_create(docType = docType, grade = grade, loanamount = "$400,000 - $499,999", purpose = purpose, maxLTV = maxLTV)
		result = FNProductMatrix.objects.get_or_create(docType = docType, grade = grade, loanamount = "$500,000 - $749,999", purpose = purpose, maxLTV = maxLTV)
		result = FNProductMatrix.objects.get_or_create(docType = docType, grade = grade, loanamount = "$750,000 - $799,999", purpose = purpose, maxLTV = maxLTV)
		result = FNProductMatrix.objects.get_or_create(docType = docType, grade = grade, loanamount = "$800,000 - $999,999", purpose = purpose, maxLTV = maxLTV)
	elif loanamount == "<=1,250,000":
		result = FNProductMatrix.objects.get_or_create(docType = docType, grade = grade, loanamount = "< $100,000", purpose = purpose, maxLTV = maxLTV)
		result = FNProductMatrix.objects.get_or_create(docType = docType, grade = grade, loanamount = "$100,000 - $149,999", purpose = purpose, maxLTV = maxLTV)
		result = FNProductMatrix.objects.get_or_create(docType = docType, grade = grade, loanamount = "$150,000 - $399,999", purpose = purpose, maxLTV = maxLTV)
		result = FNProductMatrix.objects.get_or_create(docType = docType, grade = grade, loanamount = "$400,000 - $499,999", purpose = purpose, maxLTV = maxLTV)
		result = FNProductMatrix.objects.get_or_create(docType = docType, grade = grade, loanamount = "$500,000 - $749,999", purpose = purpose, maxLTV = maxLTV)
		result = FNProductMatrix.objects.get_or_create(docType = docType, grade = grade, loanamount = "$750,000 - $799,999", purpose = purpose, maxLTV = maxLTV)
		result = FNProductMatrix.objects.get_or_create(docType = docType, grade = grade, loanamount = "$800,000 - $999,999", purpose = purpose, maxLTV = maxLTV)
		result = FNProductMatrix.objects.get_or_create(docType = docType, grade = grade, loanamount = "$1,000,000 - $1,249,999", purpose = purpose, maxLTV = maxLTV)
	elif loanamount == "1,000,001 - 1,500,000":
		result = FNProductMatrix.objects.get_or_create(docType = docType, grade = grade, loanamount = "$1,000,000 - $1,249,999", purpose = purpose, maxLTV = maxLTV)
		result = FNProductMatrix.objects.get_or_create(docType = docType, grade = grade, loanamount = "$1,250,000 - $1,499,999", purpose = purpose, maxLTV = maxLTV)
	elif loanamount == "1,500,001 - 2,000,000":
		result = FNProductMatrix.objects.get_or_create(docType = docType, grade = grade, loanamount = "$1,500,000 - $2,000,000", purpose = purpose, maxLTV = maxLTV)


def importRateSheets():
	#Reset all data models for clean refresh of rate sheets
	print "Clearing Exisiting Models..."
	CAProductMatrix.objects.all().delete()
	DTIProductMatrix.objects.all().delete()
	PIProductMatrix.objects.all().delete()
	FNProductMatrix.objects.all().delete()

	# CREDIT ASCENT PRIMARY MATRIX
	print "Loading Verus Credit Ascent Primary Matrix..."
	wb = load_workbook(filename = '/home/invictus/VerusPricer/static/matrices/Credit Ascent Primary Matrix.xlsx')
	ws = wb.active
	
	#Load Full Doc/24 Mo Bank Statement
	startRow = 6
	endRow = 24
	for row in range(startRow, endRow+1):
		if ws.cell(row=row, column=1).value is not None and ws.cell(row=row, column=3).value is not None:
			fico = ws.cell(row=row, column=3).value
			grade = ws.cell(row=row, column=1).value
			loanamount = ws.cell(row=row, column=4).value
			result = createCAModels(occupancy = "Primary", grade = grade, fico = fico, loanamount = loanamount, docType = "Full Doc", purpose = "Purchase", maxLTV = ws.cell(row=row, column=6).value)
			result = createCAModels(occupancy = "Primary", grade = grade, fico = fico, loanamount = loanamount, docType = "Full Doc", purpose = "Rate/Term Refinance", maxLTV = ws.cell(row=row, column=6).value)
			result = createCAModels(occupancy = "Primary", grade = grade, fico = fico, loanamount = loanamount, docType = "Full Doc", purpose = "Cash Out Refinance", maxLTV = ws.cell(row=row, column=7).value)
			result = createCAModels(occupancy = "Primary", grade = grade, fico = fico, loanamount = loanamount, docType = "24 Mo Bank Statement", purpose = "Purchase", maxLTV = ws.cell(row=row, column=8).value)
			result = createCAModels(occupancy = "Primary", grade = grade, fico = fico, loanamount = loanamount, docType = "24 Mo Bank Statement", purpose = "Rate/Term Refinance", maxLTV = ws.cell(row=row, column=8).value)
			result = createCAModels(occupancy = "Primary", grade = grade, fico = fico, loanamount = loanamount, docType = "24 Mo Bank Statement", purpose = "Cash Out Refinance", maxLTV = ws.cell(row=row, column=9).value)
		elif ws.cell(row=row, column=1).value is None and ws.cell(row=row, column=3).value is not None:
			fico = ws.cell(row=row, column=3).value
			loanamount = ws.cell(row=row, column=4).value
			result = createCAModels(occupancy = "Primary", grade = grade, fico = fico, loanamount = loanamount, docType = "Full Doc", purpose = "Purchase", maxLTV = ws.cell(row=row, column=6).value)
			result = createCAModels(occupancy = "Primary", grade = grade, fico = fico, loanamount = loanamount, docType = "Full Doc", purpose = "Rate/Term Refinance", maxLTV = ws.cell(row=row, column=6).value)
			result = createCAModels(occupancy = "Primary", grade = grade, fico = fico, loanamount = loanamount, docType = "Full Doc", purpose = "Cash Out Refinance", maxLTV = ws.cell(row=row, column=7).value)
			result = createCAModels(occupancy = "Primary", grade = grade, fico = fico, loanamount = loanamount, docType = "24 Mo Bank Statement", purpose = "Purchase", maxLTV = ws.cell(row=row, column=8).value)
			result = createCAModels(occupancy = "Primary", grade = grade, fico = fico, loanamount = loanamount, docType = "24 Mo Bank Statement", purpose = "Rate/Term Refinance", maxLTV = ws.cell(row=row, column=8).value)
			result = createCAModels(occupancy = "Primary", grade = grade, fico = fico, loanamount = loanamount, docType = "24 Mo Bank Statement", purpose = "Cash Out Refinance", maxLTV = ws.cell(row=row, column=9).value)
		elif ws.cell(row=row, column=1).value is None and ws.cell(row=row, column=1).value is None:
			loanamount = ws.cell(row=row, column=4).value
			result = createCAModels(occupancy = "Primary", grade = grade, fico = fico, loanamount = loanamount, docType = "Full Doc", purpose = "Purchase", maxLTV = ws.cell(row=row, column=6).value)
			result = createCAModels(occupancy = "Primary", grade = grade, fico = fico, loanamount = loanamount, docType = "Full Doc", purpose = "Rate/Term Refinance", maxLTV = ws.cell(row=row, column=6).value)
			result = createCAModels(occupancy = "Primary", grade = grade, fico = fico, loanamount = loanamount, docType = "Full Doc", purpose = "Cash Out Refinance", maxLTV = ws.cell(row=row, column=7).value)
			result = createCAModels(occupancy = "Primary", grade = grade, fico = fico, loanamount = loanamount, docType = "24 Mo Bank Statement", purpose = "Purchase", maxLTV = ws.cell(row=row, column=8).value)
			result = createCAModels(occupancy = "Primary", grade = grade, fico = fico, loanamount = loanamount, docType = "24 Mo Bank Statement", purpose = "Rate/Term Refinance", maxLTV = ws.cell(row=row, column=8).value)
			result = createCAModels(occupancy = "Primary", grade = grade, fico = fico, loanamount = loanamount, docType = "24 Mo Bank Statement", purpose = "Cash Out Refinance", maxLTV = ws.cell(row=row, column=9).value)

	#Load 12 Mo Bank Statement/Asset Utilization
	startRow = 30
	endRow = 37
	for row in range(startRow, endRow+1):
		if ws.cell(row=row, column=1).value is not None and ws.cell(row=row, column=3).value is not None:
			fico = ws.cell(row=row, column=3).value
			grade = ws.cell(row=row, column=1).value
			loanamount = ws.cell(row=row, column=4).value
			result = createCAModels(occupancy = "Primary", grade = grade, fico = fico, loanamount = loanamount, docType = "12 Mo Bank Statement", purpose = "Purchase", maxLTV = ws.cell(row=row, column=6).value)
			result = createCAModels(occupancy = "Primary", grade = grade, fico = fico, loanamount = loanamount, docType = "12 Mo Bank Statement", purpose = "Rate/Term Refinance", maxLTV = ws.cell(row=row, column=6).value)
			result = createCAModels(occupancy = "Primary", grade = grade, fico = fico, loanamount = loanamount, docType = "12 Mo Bank Statement", purpose = "Cash Out Refinance", maxLTV = ws.cell(row=row, column=7).value)
			result = createCAModels(occupancy = "Primary", grade = grade, fico = fico, loanamount = loanamount, docType = "Asset Utilization", purpose = "Purchase", maxLTV = ws.cell(row=row, column=8).value)
			result = createCAModels(occupancy = "Primary", grade = grade, fico = fico, loanamount = loanamount, docType = "Asset Utilization", purpose = "Rate/Term Refinance", maxLTV = ws.cell(row=row, column=8).value)
			result = createCAModels(occupancy = "Primary", grade = grade, fico = fico, loanamount = loanamount, docType = "Asset Utilization", purpose = "Cash Out Refinance", maxLTV = ws.cell(row=row, column=9).value)
		elif ws.cell(row=row, column=1).value is None and ws.cell(row=row, column=3).value is not None:
			fico = ws.cell(row=row, column=3).value
			loanamount = ws.cell(row=row, column=4).value
			result = createCAModels(occupancy = "Primary", grade = grade, fico = fico, loanamount = loanamount, docType = "12 Mo Bank Statement", purpose = "Purchase", maxLTV = ws.cell(row=row, column=6).value)
			result = createCAModels(occupancy = "Primary", grade = grade, fico = fico, loanamount = loanamount, docType = "12 Mo Bank Statement", purpose = "Rate/Term Refinance", maxLTV = ws.cell(row=row, column=6).value)
			result = createCAModels(occupancy = "Primary", grade = grade, fico = fico, loanamount = loanamount, docType = "12 Mo Bank Statement", purpose = "Cash Out Refinance", maxLTV = ws.cell(row=row, column=7).value)
			result = createCAModels(occupancy = "Primary", grade = grade, fico = fico, loanamount = loanamount, docType = "Asset Utilization", purpose = "Purchase", maxLTV = ws.cell(row=row, column=8).value)
			result = createCAModels(occupancy = "Primary", grade = grade, fico = fico, loanamount = loanamount, docType = "Asset Utilization", purpose = "Rate/Term Refinance", maxLTV = ws.cell(row=row, column=8).value)
			result = createCAModels(occupancy = "Primary", grade = grade, fico = fico, loanamount = loanamount, docType = "Asset Utilization", purpose = "Cash Out Refinance", maxLTV = ws.cell(row=row, column=9).value)
		elif ws.cell(row=row, column=1).value is None and ws.cell(row=row, column=1).value is None:
			loanamount = ws.cell(row=row, column=4).value
			result = createCAModels(occupancy = "Primary", grade = grade, fico = fico, loanamount = loanamount, docType = "12 Mo Bank Statement", purpose = "Purchase", maxLTV = ws.cell(row=row, column=6).value)
			result = createCAModels(occupancy = "Primary", grade = grade, fico = fico, loanamount = loanamount, docType = "12 Mo Bank Statement", purpose = "Rate/Term Refinance", maxLTV = ws.cell(row=row, column=6).value)
			result = createCAModels(occupancy = "Primary", grade = grade, fico = fico, loanamount = loanamount, docType = "12 Mo Bank Statement", purpose = "Cash Out Refinance", maxLTV = ws.cell(row=row, column=7).value)
			result = createCAModels(occupancy = "Primary", grade = grade, fico = fico, loanamount = loanamount, docType = "Asset Utilization", purpose = "Purchase", maxLTV = ws.cell(row=row, column=8).value)
			result = createCAModels(occupancy = "Primary", grade = grade, fico = fico, loanamount = loanamount, docType = "Asset Utilization", purpose = "Rate/Term Refinance", maxLTV = ws.cell(row=row, column=8).value)
			result = createCAModels(occupancy = "Primary", grade = grade, fico = fico, loanamount = loanamount, docType = "Asset Utilization", purpose = "Cash Out Refinance", maxLTV = ws.cell(row=row, column=9).value)

	# CREDIT ASCENT 2ND HOME MATRIX
	print "Loading Verus Credit Ascent 2nd Home Matrix..."
	wb = load_workbook(filename = '/home/invictus/VerusPricer/static/matrices/Credit Ascent Secondary Matrix.xlsx')
	ws = wb.active
	
	#Load Full Doc/24 Mo Bank Statement
	startRow = 6
	endRow = 23
	for row in range(startRow, endRow+1):
		if ws.cell(row=row, column=1).value is not None and ws.cell(row=row, column=3).value is not None:
			fico = ws.cell(row=row, column=3).value
			grade = ws.cell(row=row, column=1).value
			loanamount = ws.cell(row=row, column=4).value
			result = createCAModels(occupancy = "2nd Home", grade = grade, fico = fico, loanamount = loanamount, docType = "Full Doc", purpose = "Purchase", maxLTV = ws.cell(row=row, column=6).value)
			result = createCAModels(occupancy = "2nd Home", grade = grade, fico = fico, loanamount = loanamount, docType = "Full Doc", purpose = "Rate/Term Refinance", maxLTV = ws.cell(row=row, column=6).value)
			result = createCAModels(occupancy = "2nd Home", grade = grade, fico = fico, loanamount = loanamount, docType = "Full Doc", purpose = "Cash Out Refinance", maxLTV = ws.cell(row=row, column=7).value)
			result = createCAModels(occupancy = "2nd Home", grade = grade, fico = fico, loanamount = loanamount, docType = "24 Mo Bank Statement", purpose = "Purchase", maxLTV = ws.cell(row=row, column=8).value)
			result = createCAModels(occupancy = "2nd Home", grade = grade, fico = fico, loanamount = loanamount, docType = "24 Mo Bank Statement", purpose = "Rate/Term Refinance", maxLTV = ws.cell(row=row, column=8).value)
			result = createCAModels(occupancy = "2nd Home", grade = grade, fico = fico, loanamount = loanamount, docType = "24 Mo Bank Statement", purpose = "Cash Out Refinance", maxLTV = ws.cell(row=row, column=9).value)
		elif ws.cell(row=row, column=1).value is None and ws.cell(row=row, column=3).value is not None:
			fico = ws.cell(row=row, column=3).value
			loanamount = ws.cell(row=row, column=4).value
			result = createCAModels(occupancy = "2nd Home", grade = grade, fico = fico, loanamount = loanamount, docType = "Full Doc", purpose = "Purchase", maxLTV = ws.cell(row=row, column=6).value)
			result = createCAModels(occupancy = "2nd Home", grade = grade, fico = fico, loanamount = loanamount, docType = "Full Doc", purpose = "Rate/Term Refinance", maxLTV = ws.cell(row=row, column=6).value)
			result = createCAModels(occupancy = "2nd Home", grade = grade, fico = fico, loanamount = loanamount, docType = "Full Doc", purpose = "Cash Out Refinance", maxLTV = ws.cell(row=row, column=7).value)
			result = createCAModels(occupancy = "2nd Home", grade = grade, fico = fico, loanamount = loanamount, docType = "24 Mo Bank Statement", purpose = "Purchase", maxLTV = ws.cell(row=row, column=8).value)
			result = createCAModels(occupancy = "2nd Home", grade = grade, fico = fico, loanamount = loanamount, docType = "24 Mo Bank Statement", purpose = "Rate/Term Refinance", maxLTV = ws.cell(row=row, column=8).value)
			result = createCAModels(occupancy = "2nd Home", grade = grade, fico = fico, loanamount = loanamount, docType = "24 Mo Bank Statement", purpose = "Cash Out Refinance", maxLTV = ws.cell(row=row, column=9).value)
		elif ws.cell(row=row, column=1).value is None and ws.cell(row=row, column=1).value is None:
			loanamount = ws.cell(row=row, column=4).value
			result = createCAModels(occupancy = "2nd Home", grade = grade, fico = fico, loanamount = loanamount, docType = "Full Doc", purpose = "Purchase", maxLTV = ws.cell(row=row, column=6).value)
			result = createCAModels(occupancy = "2nd Home", grade = grade, fico = fico, loanamount = loanamount, docType = "Full Doc", purpose = "Rate/Term Refinance", maxLTV = ws.cell(row=row, column=6).value)
			result = createCAModels(occupancy = "2nd Home", grade = grade, fico = fico, loanamount = loanamount, docType = "Full Doc", purpose = "Cash Out Refinance", maxLTV = ws.cell(row=row, column=7).value)
			result = createCAModels(occupancy = "2nd Home", grade = grade, fico = fico, loanamount = loanamount, docType = "24 Mo Bank Statement", purpose = "Purchase", maxLTV = ws.cell(row=row, column=8).value)
			result = createCAModels(occupancy = "2nd Home", grade = grade, fico = fico, loanamount = loanamount, docType = "24 Mo Bank Statement", purpose = "Rate/Term Refinance", maxLTV = ws.cell(row=row, column=8).value)
			result = createCAModels(occupancy = "2nd Home", grade = grade, fico = fico, loanamount = loanamount, docType = "24 Mo Bank Statement", purpose = "Cash Out Refinance", maxLTV = ws.cell(row=row, column=9).value)

	#Load 12 Mo Bank Statement/Asset Utilization
	startRow = 29
	endRow = 32
	for row in range(startRow, endRow+1):
		if ws.cell(row=row, column=1).value is not None and ws.cell(row=row, column=3).value is not None:
			fico = ws.cell(row=row, column=3).value
			grade = ws.cell(row=row, column=1).value
			loanamount = ws.cell(row=row, column=4).value
			result = createCAModels(occupancy = "2nd Home", grade = grade, fico = fico, loanamount = loanamount, docType = "12 Mo Bank Statement", purpose = "Purchase", maxLTV = ws.cell(row=row, column=6).value)
			result = createCAModels(occupancy = "2nd Home", grade = grade, fico = fico, loanamount = loanamount, docType = "12 Mo Bank Statement", purpose = "Rate/Term Refinance", maxLTV = ws.cell(row=row, column=6).value)
			result = createCAModels(occupancy = "2nd Home", grade = grade, fico = fico, loanamount = loanamount, docType = "12 Mo Bank Statement", purpose = "Cash Out Refinance", maxLTV = ws.cell(row=row, column=7).value)
			result = createCAModels(occupancy = "2nd Home", grade = grade, fico = fico, loanamount = loanamount, docType = "Asset Utilization", purpose = "Purchase", maxLTV = ws.cell(row=row, column=8).value)
			result = createCAModels(occupancy = "2nd Home", grade = grade, fico = fico, loanamount = loanamount, docType = "Asset Utilization", purpose = "Rate/Term Refinance", maxLTV = ws.cell(row=row, column=8).value)
			result = createCAModels(occupancy = "2nd Home", grade = grade, fico = fico, loanamount = loanamount, docType = "Asset Utilization", purpose = "Cash Out Refinance", maxLTV = ws.cell(row=row, column=9).value)
		elif ws.cell(row=row, column=1).value is None and ws.cell(row=row, column=3).value is not None:
			fico = ws.cell(row=row, column=3).value
			loanamount = ws.cell(row=row, column=4).value
			result = createCAModels(occupancy = "2nd Home", grade = grade, fico = fico, loanamount = loanamount, docType = "12 Mo Bank Statement", purpose = "Purchase", maxLTV = ws.cell(row=row, column=6).value)
			result = createCAModels(occupancy = "2nd Home", grade = grade, fico = fico, loanamount = loanamount, docType = "12 Mo Bank Statement", purpose = "Rate/Term Refinance", maxLTV = ws.cell(row=row, column=6).value)
			result = createCAModels(occupancy = "2nd Home", grade = grade, fico = fico, loanamount = loanamount, docType = "12 Mo Bank Statement", purpose = "Cash Out Refinance", maxLTV = ws.cell(row=row, column=7).value)
			result = createCAModels(occupancy = "2nd Home", grade = grade, fico = fico, loanamount = loanamount, docType = "Asset Utilization", purpose = "Purchase", maxLTV = ws.cell(row=row, column=8).value)
			result = createCAModels(occupancy = "2nd Home", grade = grade, fico = fico, loanamount = loanamount, docType = "Asset Utilization", purpose = "Rate/Term Refinance", maxLTV = ws.cell(row=row, column=8).value)
			result = createCAModels(occupancy = "2nd Home", grade = grade, fico = fico, loanamount = loanamount, docType = "Asset Utilization", purpose = "Cash Out Refinance", maxLTV = ws.cell(row=row, column=9).value)
		elif ws.cell(row=row, column=1).value is None and ws.cell(row=row, column=1).value is None:
			loanamount = ws.cell(row=row, column=4).value
			result = createCAModels(occupancy = "2nd Home", grade = grade, fico = fico, loanamount = loanamount, docType = "12 Mo Bank Statement", purpose = "Purchase", maxLTV = ws.cell(row=row, column=6).value)
			result = createCAModels(occupancy = "2nd Home", grade = grade, fico = fico, loanamount = loanamount, docType = "12 Mo Bank Statement", purpose = "Rate/Term Refinance", maxLTV = ws.cell(row=row, column=6).value)
			result = createCAModels(occupancy = "2nd Home", grade = grade, fico = fico, loanamount = loanamount, docType = "12 Mo Bank Statement", purpose = "Cash Out Refinance", maxLTV = ws.cell(row=row, column=7).value)
			result = createCAModels(occupancy = "2nd Home", grade = grade, fico = fico, loanamount = loanamount, docType = "Asset Utilization", purpose = "Purchase", maxLTV = ws.cell(row=row, column=8).value)
			result = createCAModels(occupancy = "2nd Home", grade = grade, fico = fico, loanamount = loanamount, docType = "Asset Utilization", purpose = "Rate/Term Refinance", maxLTV = ws.cell(row=row, column=8).value)
			result = createCAModels(occupancy = "2nd Home", grade = grade, fico = fico, loanamount = loanamount, docType = "Asset Utilization", purpose = "Cash Out Refinance", maxLTV = ws.cell(row=row, column=9).value)

	# INVESTOR DTI MATRIX
	print "Loading Verus Investor DTI Matrix..."
	wb = load_workbook(filename = '/home/invictus/VerusPricer/static/matrices/Investor DTI Matrix.xlsx')
	ws = wb.active
	
	#Load Full Doc/24 Mo Bank Statement
	startRow = 6
	endRow = 18
	for row in range(startRow, endRow+1):
		if ws.cell(row=row, column=1).value is not None and ws.cell(row=row, column=3).value is not None:
			fico = ws.cell(row=row, column=3).value
			grade = ws.cell(row=row, column=1).value
			loanamount = ws.cell(row=row, column=4).value
			result = createDTIModels(grade = grade, fico = fico, loanamount = loanamount, docType = "Full Doc", purpose = "Purchase", maxLTV = ws.cell(row=row, column=6).value)
			result = createDTIModels(grade = grade, fico = fico, loanamount = loanamount, docType = "Full Doc", purpose = "Rate/Term Refinance", maxLTV = ws.cell(row=row, column=6).value)
			result = createDTIModels(grade = grade, fico = fico, loanamount = loanamount, docType = "Full Doc", purpose = "Cash Out Refinance", maxLTV = ws.cell(row=row, column=7).value)
			result = createDTIModels(grade = grade, fico = fico, loanamount = loanamount, docType = "24 Mo Bank Statement", purpose = "Purchase", maxLTV = ws.cell(row=row, column=8).value)
			result = createDTIModels(grade = grade, fico = fico, loanamount = loanamount, docType = "24 Mo Bank Statement", purpose = "Rate/Term Refinance", maxLTV = ws.cell(row=row, column=8).value)
			result = createDTIModels(grade = grade, fico = fico, loanamount = loanamount, docType = "24 Mo Bank Statement", purpose = "Cash Out Refinance", maxLTV = ws.cell(row=row, column=9).value)
		elif ws.cell(row=row, column=1).value is None and ws.cell(row=row, column=3).value is not None:
			fico = ws.cell(row=row, column=3).value
			loanamount = ws.cell(row=row, column=4).value
			result = createDTIModels(grade = grade, fico = fico, loanamount = loanamount, docType = "Full Doc", purpose = "Purchase", maxLTV = ws.cell(row=row, column=6).value)
			result = createDTIModels(grade = grade, fico = fico, loanamount = loanamount, docType = "Full Doc", purpose = "Rate/Term Refinance", maxLTV = ws.cell(row=row, column=6).value)
			result = createDTIModels(grade = grade, fico = fico, loanamount = loanamount, docType = "Full Doc", purpose = "Cash Out Refinance", maxLTV = ws.cell(row=row, column=7).value)
			result = createDTIModels(grade = grade, fico = fico, loanamount = loanamount, docType = "24 Mo Bank Statement", purpose = "Purchase", maxLTV = ws.cell(row=row, column=8).value)
			result = createDTIModels(grade = grade, fico = fico, loanamount = loanamount, docType = "24 Mo Bank Statement", purpose = "Rate/Term Refinance", maxLTV = ws.cell(row=row, column=8).value)
			result = createDTIModels(grade = grade, fico = fico, loanamount = loanamount, docType = "24 Mo Bank Statement", purpose = "Cash Out Refinance", maxLTV = ws.cell(row=row, column=9).value)
		elif ws.cell(row=row, column=1).value is None and ws.cell(row=row, column=1).value is None:
			loanamount = ws.cell(row=row, column=4).value
			result = createDTIModels(grade = grade, fico = fico, loanamount = loanamount, docType = "Full Doc", purpose = "Purchase", maxLTV = ws.cell(row=row, column=6).value)
			result = createDTIModels(grade = grade, fico = fico, loanamount = loanamount, docType = "Full Doc", purpose = "Rate/Term Refinance", maxLTV = ws.cell(row=row, column=6).value)
			result = createDTIModels(grade = grade, fico = fico, loanamount = loanamount, docType = "Full Doc", purpose = "Cash Out Refinance", maxLTV = ws.cell(row=row, column=7).value)
			result = createDTIModels(grade = grade, fico = fico, loanamount = loanamount, docType = "24 Mo Bank Statement", purpose = "Purchase", maxLTV = ws.cell(row=row, column=8).value)
			result = createDTIModels(grade = grade, fico = fico, loanamount = loanamount, docType = "24 Mo Bank Statement", purpose = "Rate/Term Refinance", maxLTV = ws.cell(row=row, column=8).value)
			result = createDTIModels(grade = grade, fico = fico, loanamount = loanamount, docType = "24 Mo Bank Statement", purpose = "Cash Out Refinance", maxLTV = ws.cell(row=row, column=9).value)

	#Load 12 Mo Bank Statement/Asset Utilization
	startRow = 24
	endRow = 27
	for row in range(startRow, endRow+1):
		if ws.cell(row=row, column=1).value is not None and ws.cell(row=row, column=3).value is not None:
			fico = ws.cell(row=row, column=3).value
			grade = ws.cell(row=row, column=1).value
			loanamount = ws.cell(row=row, column=4).value
			result = createDTIModels(grade = grade, fico = fico, loanamount = loanamount, docType = "12 Mo Bank Statement", purpose = "Purchase", maxLTV = ws.cell(row=row, column=6).value)
			result = createDTIModels(grade = grade, fico = fico, loanamount = loanamount, docType = "12 Mo Bank Statement", purpose = "Rate/Term Refinance", maxLTV = ws.cell(row=row, column=6).value)
			result = createDTIModels(grade = grade, fico = fico, loanamount = loanamount, docType = "12 Mo Bank Statement", purpose = "Cash Out Refinance", maxLTV = ws.cell(row=row, column=7).value)
			result = createDTIModels(grade = grade, fico = fico, loanamount = loanamount, docType = "Asset Utilization", purpose = "Purchase", maxLTV = ws.cell(row=row, column=8).value)
			result = createDTIModels(grade = grade, fico = fico, loanamount = loanamount, docType = "Asset Utilization", purpose = "Rate/Term Refinance", maxLTV = ws.cell(row=row, column=8).value)
			result = createDTIModels(grade = grade, fico = fico, loanamount = loanamount, docType = "Asset Utilization", purpose = "Cash Out Refinance", maxLTV = ws.cell(row=row, column=9).value)
		elif ws.cell(row=row, column=1).value is None and ws.cell(row=row, column=3).value is not None:
			fico = ws.cell(row=row, column=3).value
			loanamount = ws.cell(row=row, column=4).value
			result = createDTIModels(grade = grade, fico = fico, loanamount = loanamount, docType = "12 Mo Bank Statement", purpose = "Purchase", maxLTV = ws.cell(row=row, column=6).value)
			result = createDTIModels(grade = grade, fico = fico, loanamount = loanamount, docType = "12 Mo Bank Statement", purpose = "Rate/Term Refinance", maxLTV = ws.cell(row=row, column=6).value)
			result = createDTIModels(grade = grade, fico = fico, loanamount = loanamount, docType = "12 Mo Bank Statement", purpose = "Cash Out Refinance", maxLTV = ws.cell(row=row, column=7).value)
			result = createDTIModels(grade = grade, fico = fico, loanamount = loanamount, docType = "Asset Utilization", purpose = "Purchase", maxLTV = ws.cell(row=row, column=8).value)
			result = createDTIModels(grade = grade, fico = fico, loanamount = loanamount, docType = "Asset Utilization", purpose = "Rate/Term Refinance", maxLTV = ws.cell(row=row, column=8).value)
			result = createDTIModels(grade = grade, fico = fico, loanamount = loanamount, docType = "Asset Utilization", purpose = "Cash Out Refinance", maxLTV = ws.cell(row=row, column=9).value)
		elif ws.cell(row=row, column=1).value is None and ws.cell(row=row, column=1).value is None:
			loanamount = ws.cell(row=row, column=4).value
			result = createDTIModels(grade = grade, fico = fico, loanamount = loanamount, docType = "12 Mo Bank Statement", purpose = "Purchase", maxLTV = ws.cell(row=row, column=6).value)
			result = createDTIModels(grade = grade, fico = fico, loanamount = loanamount, docType = "12 Mo Bank Statement", purpose = "Rate/Term Refinance", maxLTV = ws.cell(row=row, column=6).value)
			result = createDTIModels(grade = grade, fico = fico, loanamount = loanamount, docType = "12 Mo Bank Statement", purpose = "Cash Out Refinance", maxLTV = ws.cell(row=row, column=7).value)
			result = createDTIModels(grade = grade, fico = fico, loanamount = loanamount, docType = "Asset Utilization", purpose = "Purchase", maxLTV = ws.cell(row=row, column=8).value)
			result = createDTIModels(grade = grade, fico = fico, loanamount = loanamount, docType = "Asset Utilization", purpose = "Rate/Term Refinance", maxLTV = ws.cell(row=row, column=8).value)
			result = createDTIModels(grade = grade, fico = fico, loanamount = loanamount, docType = "Asset Utilization", purpose = "Cash Out Refinance", maxLTV = ws.cell(row=row, column=9).value)

 # INVESTOR PI MATRIX
 	print "Loading Verus Investor PI Matrix..."
 	wb = load_workbook(filename = '/home/maverick/VerusPricer/static/matrices/Investor PI Matrix.xlsx')
 	ws = wb.active
	
 	#Load DSCR/No Ratio Matrices
 	startRow = 4
 	endRow = 27
 	for row in range(startRow, endRow+1):
 		if ws.cell(row=row, column=3).value is not None and ws.cell(row=row, column=4).value is not None and ws.cell(row=row, column=5).value is not None:
 			fico = ws.cell(row=row, column=4).value
 			doctype = ws.cell(row=row, column=3).value
 			loanamount = ws.cell(row=row, column=5).value
 			reserves = ws.cell(row=row, column=7).value.replace(" Months", "")
 			result = createPIModels(docType = doctype, fico = fico, loanamount = loanamount, reserves = reserves, purpose = "Purchase", maxLTV = ws.cell(row=row, column=8).value)
 			result = createPIModels(docType = doctype, fico = fico, loanamount = loanamount, reserves = reserves, purpose = "Rate/Term Refinance", maxLTV = ws.cell(row=row, column=9).value)
 			result = createPIModels(docType = doctype, fico = fico, loanamount = loanamount, reserves = reserves, purpose = "Cash Out Refinance", maxLTV = ws.cell(row=row, column=10).value)
 		elif ws.cell(row=row, column=3).value is None and ws.cell(row=row, column=4).value is None and ws.cell(row=row, column=5).value is not None:
 			loanamount = ws.cell(row=row, column=5).value
 			reserves = ws.cell(row=row, column=7).value.replace(" Months", "")
 			result = createPIModels(docType = doctype, fico = fico, loanamount = loanamount, reserves = reserves, purpose = "Purchase", maxLTV = ws.cell(row=row, column=8).value)
 			result = createPIModels(docType = doctype, fico = fico, loanamount = loanamount, reserves = reserves, purpose = "Rate/Term Refinance", maxLTV = ws.cell(row=row, column=9).value)
 			result = createPIModels(docType = doctype, fico = fico, loanamount = loanamount, reserves = reserves, purpose = "Cash Out Refinance", maxLTV = ws.cell(row=row, column=10).value)
 		elif ws.cell(row=row, column=3).value is None and ws.cell(row=row, column=4).value is not None:
 			fico = ws.cell(row=row, column=4).value
 			loanamount = ws.cell(row=row, column=5).value
 			reserves = ws.cell(row=row, column=7).value.replace(" Months", "")
 			result = createPIModels(docType = doctype, fico = fico, loanamount = loanamount, reserves = reserves, purpose = "Purchase", maxLTV = ws.cell(row=row, column=8).value)
 			result = createPIModels(docType = doctype, fico = fico, loanamount = loanamount, reserves = reserves, purpose = "Rate/Term Refinance", maxLTV = ws.cell(row=row, column=9).value)
 			result = createPIModels(docType = doctype, fico = fico, loanamount = loanamount, reserves = reserves, purpose = "Cash Out Refinance", maxLTV = ws.cell(row=row, column=10).value)
 		elif ws.cell(row=row, column=3).value is None and ws.cell(row=row, column=4).value is None and ws.cell(row=row, column=5).value is None:
 			reserves = ws.cell(row=row, column=7).value.replace(" Months", "")
 			result = createPIModels(docType = doctype, fico = fico, loanamount = loanamount, reserves = reserves, purpose = "Purchase", maxLTV = ws.cell(row=row, column=8).value)
 			result = createPIModels(docType = doctype, fico = fico, loanamount = loanamount, reserves = reserves, purpose = "Rate/Term Refinance", maxLTV = ws.cell(row=row, column=9).value)
 			result = createPIModels(docType = doctype, fico = fico, loanamount = loanamount, reserves = reserves, purpose = "Cash Out Refinance", maxLTV = ws.cell(row=row, column=10).value)

 # FOREIGN NATIONAL MATRIX
 	print "Loading Verus Foreign National Matrix..."
 	wb = load_workbook(filename = '/home/maverick/VerusPricer/static/matrices/Foreign National Matrix.xlsx')
 	ws = wb.active
	
 	#Load Full Doc Matrices
 	startRow = 3
 	endRow = 8
 	for row in range(startRow, endRow+1):
 		if ws.cell(row=row, column=1).value is not None and ws.cell(row=row, column=2).value is not None:
 			doctype = ws.cell(row=row, column=1).value
 			grade = ws.cell(row=row, column=2).value
 			loanamount = ws.cell(row=row, column=3).value
 			result = createFNModels(docType = doctype, grade = grade, loanamount = loanamount, purpose = "Purchase", maxLTV = ws.cell(row=row, column=4).value)
 			result = createFNModels(docType = doctype, grade = grade, loanamount = loanamount, purpose = "Rate/Term Refinance", maxLTV = ws.cell(row=row, column=5).value)
 			result = createFNModels(docType = doctype, grade = grade, loanamount = loanamount, purpose = "Cash Out Refinance", maxLTV = ws.cell(row=row, column=6).value)
 		elif ws.cell(row=row, column=1).value is None and ws.cell(row=row, column=2).value is not None:
 			grade = ws.cell(row=row, column=2).value
 			loanamount = ws.cell(row=row, column=3).value
 			result = createFNModels(docType = doctype, grade = grade, loanamount = loanamount, purpose = "Purchase", maxLTV = ws.cell(row=row, column=4).value)
 			result = createFNModels(docType = doctype, grade = grade, loanamount = loanamount, purpose = "Rate/Term Refinance", maxLTV = ws.cell(row=row, column=5).value)
 			result = createFNModels(docType = doctype, grade = grade, loanamount = loanamount, purpose = "Cash Out Refinance", maxLTV = ws.cell(row=row, column=6).value)
 		elif ws.cell(row=row, column=1).value is None and ws.cell(row=row, column=2).value is None:
 			loanamount = ws.cell(row=row, column=3).value
 			result = createFNModels(docType = doctype, grade = grade, loanamount = loanamount, purpose = "Purchase", maxLTV = ws.cell(row=row, column=4).value)
 			result = createFNModels(docType = doctype, grade = grade, loanamount = loanamount, purpose = "Rate/Term Refinance", maxLTV = ws.cell(row=row, column=5).value)
 			result = createFNModels(docType = doctype, grade = grade, loanamount = loanamount, purpose = "Cash Out Refinance", maxLTV = ws.cell(row=row, column=6).value)

 	#Load DSCR/No Ratio Matrices
 	startRow = 9
 	endRow = 12
 	for row in range(startRow, endRow+1):
 		if ws.cell(row=row, column=1).value is not None:
 			doctype = ws.cell(row=row, column=1).value
 			grade = "N/A"
 			loanamount = ws.cell(row=row, column=3).value
 			result = createFNModels(docType = doctype, grade = grade, loanamount = loanamount, purpose = "Purchase", maxLTV = ws.cell(row=row, column=4).value)
 			result = createFNModels(docType = doctype, grade = grade, loanamount = loanamount, purpose = "Rate/Term Refinance", maxLTV = ws.cell(row=row, column=5).value)
 			result = createFNModels(docType = doctype, grade = grade, loanamount = loanamount, purpose = "Cash Out Refinance", maxLTV = ws.cell(row=row, column=6).value)
 		elif ws.cell(row=row, column=1).value is None and ws.cell(row=row, column=2).value is None:
 			loanamount = ws.cell(row=row, column=3).value
 			result = createFNModels(docType = doctype, grade = grade, loanamount = loanamount, purpose = "Purchase", maxLTV = ws.cell(row=row, column=4).value)
 			result = createFNModels(docType = doctype, grade = grade, loanamount = loanamount, purpose = "Rate/Term Refinance", maxLTV = ws.cell(row=row, column=5).value)
 			result = createFNModels(docType = doctype, grade = grade, loanamount = loanamount, purpose = "Cash Out Refinance", maxLTV = ws.cell(row=row, column=6).value)

if __name__ == '__main__':
	print "Importing Rate Sheets..."
	importRateSheets()
